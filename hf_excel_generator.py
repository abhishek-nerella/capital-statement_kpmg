"""
Hedge Fund PCAP — 10-sheet openpyxl workbook generator.

Input:
  pcap_df      — clean DataFrame from hf_pcap_engine.read_hf_pcap()
  cf_ledger_df — optional transaction-level cashflow DataFrame

Sheets:
  1. Period_Params          — fund parameters (single source of truth)
  2. Dashboard              — fund KPI summary + investor breakdown
  3. Investor_Register      — per-investor static / compliance data
  4. PCAP                   — full raw PCAP (all columns)
  5. Capital_Accounts       — simplified 9-line per-investor CQ/YTD/ITD
  6. CF_Ledger              — transaction-level cashflow (optional input)
  7. CF_Aggregator          — investor × quarter aggregation
  8. Distribution_Waterfall — waterfall analysis per investor
  9. Cashflow_IRR           — IRR cashflow vectors + XIRR formula
 10. Stress_IRR             — NAV haircut scenarios (0%, −10%, −20%, −30%, custom)
"""

from __future__ import annotations

import io
import math
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── KPMG colour palette ────────────────────────────────────────────────────────
_BLUE  = "00338D"
_LIGHT = "ACEAFF"
_NAVY  = "0C233C"
_WHITE = "FFFFFF"
_GRAY1 = "F2F6FF"
_GRAY2 = "FFFFFF"
_TOTAL = "E8F4FF"
_SEC   = "D0E8FF"
_WARN  = "FFF2CC"


# ── Style factories ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=9, color=_NAVY, italic=False) -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(color=_LIGHT) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row: int, col: int, value: str, width: float | None = None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=True, color=_WHITE, size=9)
    c.fill      = _fill(_BLUE)
    c.alignment = _align("center", wrap=True)
    c.border    = _border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _cell(ws, row: int, col: int, value: Any = None,
          fmt: str | None = None, bold: bool = False,
          align: str = "right", fill_hex: str | None = None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, size=9)
    c.alignment = _align(align)
    c.border    = _border()
    if fmt:
        c.number_format = fmt
    if fill_hex:
        c.fill = _fill(fill_hex)


def _banner(ws, title: str, subtitle: str = "", n_cols: int = 12) -> int:
    last = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]
    c.value     = title
    c.font      = _font(bold=True, size=13, color=_WHITE)
    c.fill      = _fill(_BLUE)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 26
    if subtitle:
        ws.merge_cells(f"A2:{last}2")
        c2 = ws["A2"]
        c2.value     = subtitle
        c2.font      = _font(size=9, italic=True)
        c2.fill      = _fill(_LIGHT)
        c2.alignment = _align("center")
        return 3
    return 2


def _section_banner(ws, row: int, text: str, n_cols: int) -> None:
    end_col = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{end_col}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(bold=True, color=_NAVY, size=10)
    c.fill      = _fill(_SEC)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 16


def _inv_banner(ws, row: int, investor: str, n_cols: int) -> None:
    end_col = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{end_col}{row}")
    c = ws.cell(row=row, column=1, value=f"Investor: {investor}")
    c.font      = _font(bold=True, color=_WHITE, size=10)
    c.fill      = _fill(_NAVY)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 18


def _placeholder(ws, row: int, message: str, n_cols: int = 8) -> None:
    end_col = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{end_col}{row}")
    c = ws.cell(row=row, column=1, value=message)
    c.font      = _font(italic=True, color="888888", size=10)
    c.fill      = _fill(_WARN)
    c.alignment = _align("center")


# ── Number formats ─────────────────────────────────────────────────────────────
_USD   = '"$"#,##0.00'
_PCT_R = '0.00"%"'
_UNITS = '#,##0.0000'
_PX    = '"$"#,##0.0000'
_X     = '0.00"x"'
_INT   = '#,##0'
_DATEFMT = 'YYYY-MM-DD'


# ── Safe accessors ─────────────────────────────────────────────────────────────

def _gv(row: pd.Series, field: str, default=None):
    val = row.get(field, default)
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    return val


def _gf(row: pd.Series, field: str, default: float = 0.0) -> float:
    val = _gv(row, field, default)
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _gt(row: pd.Series, field: str, default: str = "—") -> str:
    val = _gv(row, field, "")
    s = str(val).strip() if val is not None else ""
    return s if s and s not in ("nan", "None", "") else default


def _col_sum(pcap: pd.DataFrame, col: str) -> float:
    if col in pcap.columns:
        return float(pcap[col].dropna().sum())
    return 0.0


def _col_mean(pcap: pd.DataFrame, col: str) -> float:
    if col in pcap.columns:
        s = pcap[col].dropna()
        return float(s.mean()) if len(s) > 0 else 0.0
    return 0.0


def _col_first(pcap: pd.DataFrame, col: str, default=None):
    if col in pcap.columns:
        vals = pcap[col].dropna()
        if len(vals) > 0:
            return vals.iloc[0]
    return default


def _parse_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None:
        return None
    try:
        return datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        pass
    try:
        return datetime.strptime(str(v).strip(), "%m/%d/%Y").date()
    except Exception:
        return None


def _stress_irr(stressed_nav: float, dist_itd: float, contrib_itd: float,
                hold_days: int) -> float:
    if contrib_itd <= 0 or hold_days <= 0:
        return 0.0
    terminal = stressed_nav + dist_itd
    if terminal <= 0:
        return -1.0
    try:
        return (terminal / contrib_itd) ** (365.0 / hold_days) - 1.0
    except Exception:
        return 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1: Period_Params
# ═══════════════════════════════════════════════════════════════════════════════

def _build_period_params(ws, pcap: pd.DataFrame) -> None:
    ws.title = "Period_Params"
    r = _banner(ws, "Period Parameters — Single Source of Truth",
                "Fund-level parameters derived from pre-calculated PCAP data", n_cols=6)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 26

    _section_banner(ws, r, "A.  Fund-Level Parameters", 3)
    r += 1
    _hdr(ws, r, 1, "Parameter")
    _hdr(ws, r, 2, "Value")
    _hdr(ws, r, 3, "Notes")
    r += 1

    beg_px  = _col_first(pcap, "BEG_PX",  0.0)
    end_px  = _col_first(pcap, "END_PX",  0.0)

    params: list[tuple] = [
        ("Number of Investors",           int(len(pcap)),                      None,    "Integer count of LP investors"),
        ("Total AUM — Ending NAV (CQ)",   _col_sum(pcap, "END_CAP_CQ"),        _USD,    "Sum of all ending partner's capital (CQ)"),
        ("Total AUM — Ending NAV (ITD)",  _col_sum(pcap, "END_CAP_ITD"),       _USD,    "Sum of all ending partner's capital (ITD)"),
        ("Total Contributions (ITD)",     _col_sum(pcap, "CONTRIB_ITD"),        _USD,    "Sum of all LP capital contributions"),
        ("Total LP Distributions (ITD)",  _col_sum(pcap, "DIST_LP_ITD"),       _USD,    "Sum of all distributions to LPs"),
        ("Total Incentive Fees (ITD)",    _col_sum(pcap, "INC_FEE_ITD"),        _USD,    "Sum of all incentive/performance fees"),
        ("Total Mgmt. Fees (ITD)",        _col_sum(pcap, "MGMT_FEE_ITD_DLR"),  _USD,    "Sum of all management fees (dollar)"),
        ("Opening Unit Price",            float(beg_px) if beg_px else 0.0,    _PX,     "Beginning of period NAV per unit"),
        ("Ending Unit Price",             float(end_px) if end_px else 0.0,    _PX,     "End of period NAV per unit"),
        ("Unit Price Change",             (float(end_px) - float(beg_px)) if (beg_px and end_px) else 0.0, _PX, "Ending − Opening unit price"),
        ("Average Gross IRR",             _col_mean(pcap, "GROSS_IRR"),         _PCT_R,  "Equal-weighted average across all investors"),
        ("Average Net IRR",               _col_mean(pcap, "NET_IRR"),           _PCT_R,  "Equal-weighted average across all investors"),
        ("Average TVPI",                  _col_mean(pcap, "TVPI"),              _X,      "Equal-weighted average"),
        ("Average DPI",                   _col_mean(pcap, "DPI"),               _X,      "Equal-weighted average"),
        ("Average RVPI",                  _col_mean(pcap, "RVPI"),              _X,      "Equal-weighted average"),
        ("LP Net Waterfall (Total ITD)",  _col_sum(pcap, "LP_NET_WF"),          _USD,    "Sum of all LP net waterfall shares"),
    ]

    for i, (label, val, fmt, notes) in enumerate(params):
        fill = _GRAY1 if i % 2 else _GRAY2
        row_idx = r + i
        _cell(ws, row_idx, 1, label, align="left",  fill_hex=fill)
        _cell(ws, row_idx, 2, val,   fmt=fmt, align="right" if fmt else "left", fill_hex=fill)
        _cell(ws, row_idx, 3, notes, align="left",  fill_hex=fill)
    r += len(params)

    # Fee structures
    r += 1
    _section_banner(ws, r, "B.  Fee Structures by Investor", 6)
    r += 1
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    _hdr(ws, r, 1, "Investor",       width=36)
    _hdr(ws, r, 2, "Mgmt. Fee %",    width=16)
    _hdr(ws, r, 3, "Incentive Fee %",width=18)
    _hdr(ws, r, 4, "Pref. Return %", width=16)
    _hdr(ws, r, 5, "Hurdle Type",    width=16)
    _hdr(ws, r, 6, "Catch-up %",     width=14)
    r += 1

    for i, (_, row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2
        _cell(ws, r, 1, _gt(row, "INVESTOR_NAME"), align="left",  fill_hex=fill)
        _cell(ws, r, 2, _gf(row, "MGMT_FEE_RATE"), fmt=_PCT_R,   fill_hex=fill)
        _cell(ws, r, 3, _gf(row, "INC_FEE_RATE"),  fmt=_PCT_R,   fill_hex=fill)
        _cell(ws, r, 4, _gf(row, "PREF_RET"),      fmt=_PCT_R,   fill_hex=fill)
        _cell(ws, r, 5, _gt(row, "HURDLE_TYPE"),   align="left",  fill_hex=fill)
        _cell(ws, r, 6, _gf(row, "CATCHUP_PCT"),   fmt=_PCT_R,   fill_hex=fill)
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2: Dashboard
# ═══════════════════════════════════════════════════════════════════════════════

def _build_dashboard(ws, pcap: pd.DataFrame) -> None:
    ws.title = "Dashboard"
    r = _banner(ws, "Fund Dashboard — KPI Summary", n_cols=9)

    kpis = [
        ("Total Investors",               int(len(pcap)),                   None,   "left"),
        ("Total AUM (Ending NAV, CQ)",    _col_sum(pcap, "END_CAP_CQ"),     _USD,   "right"),
        ("Total Contributions (ITD)",     _col_sum(pcap, "CONTRIB_ITD"),    _USD,   "right"),
        ("Total LP Distributions (ITD)",  _col_sum(pcap, "DIST_LP_ITD"),   _USD,   "right"),
        ("Total Incentive Fees (ITD)",    _col_sum(pcap, "INC_FEE_ITD"),   _USD,   "right"),
        ("LP Net Waterfall (ITD)",        _col_sum(pcap, "LP_NET_WF"),     _USD,   "right"),
        ("Average Gross IRR",             _col_mean(pcap, "GROSS_IRR"),    _PCT_R, "right"),
        ("Average Net IRR",               _col_mean(pcap, "NET_IRR"),      _PCT_R, "right"),
        ("Average TVPI",                  _col_mean(pcap, "TVPI"),         _X,     "right"),
        ("Ending Unit Price",             _col_first(pcap, "END_PX", 0.0), _PX,    "right"),
    ]

    _hdr(ws, r, 1, "KPI",   width=34)
    _hdr(ws, r, 2, "Value", width=22)
    r += 1

    for i, (label, val, fmt, align) in enumerate(kpis):
        fill = _GRAY1 if i % 2 else _GRAY2
        _cell(ws, r, 1, label, align="left",  fill_hex=fill)
        _cell(ws, r, 2, val,   fmt=fmt, align=align, fill_hex=fill)
        r += 1

    # Investor breakdown table
    r += 1
    _section_banner(ws, r, "Investor Breakdown", 9)
    r += 1

    hdrs   = ["Investor", "Inception", "Currency", "NAV (CQ)", "NAV (ITD)",
              "Gross IRR", "Net IRR", "TVPI", "LP Net WF"]
    widths = [26, 14, 10, 18, 18, 12, 12, 10, 18]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws, r, ci, h, width=w)
    r += 1

    for i, (_, row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2
        vals = [
            (_gt(row, "INVESTOR_NAME"),  "left",  None),
            (_gt(row, "INCEPTION_DATE"), "left",  None),
            (_gt(row, "REPT_CCY"),       "left",  None),
            (_gf(row, "END_CAP_CQ"),    "right", _USD),
            (_gf(row, "END_CAP_ITD"),   "right", _USD),
            (_gf(row, "GROSS_IRR"),     "right", _PCT_R),
            (_gf(row, "NET_IRR"),       "right", _PCT_R),
            (_gf(row, "TVPI"),          "right", _X),
            (_gf(row, "LP_NET_WF"),     "right", _USD),
        ]
        for ci, (v, align, fmt) in enumerate(vals, 1):
            _cell(ws, r, ci, v, fmt=fmt, align=align, fill_hex=fill)
        r += 1

    # Totals row
    fill = _TOTAL
    _cell(ws, r, 1, "TOTAL / AVERAGE", bold=True, align="left", fill_hex=fill)
    _cell(ws, r, 2, "",   align="left",  fill_hex=fill)
    _cell(ws, r, 3, "",   align="left",  fill_hex=fill)
    _cell(ws, r, 4, _col_sum(pcap, "END_CAP_CQ"),  fmt=_USD,   bold=True, fill_hex=fill)
    _cell(ws, r, 5, _col_sum(pcap, "END_CAP_ITD"), fmt=_USD,   bold=True, fill_hex=fill)
    _cell(ws, r, 6, _col_mean(pcap, "GROSS_IRR"),  fmt=_PCT_R, bold=True, fill_hex=fill)
    _cell(ws, r, 7, _col_mean(pcap, "NET_IRR"),    fmt=_PCT_R, bold=True, fill_hex=fill)
    _cell(ws, r, 8, _col_mean(pcap, "TVPI"),       fmt=_X,     bold=True, fill_hex=fill)
    _cell(ws, r, 9, _col_sum(pcap, "LP_NET_WF"),   fmt=_USD,   bold=True, fill_hex=fill)


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3: Investor_Register
# ═══════════════════════════════════════════════════════════════════════════════

def _build_investor_register(ws, pcap: pd.DataFrame) -> None:
    ws.title = "Investor_Register"
    r = _banner(ws, "Investor Register — Static & Compliance Data", n_cols=16)

    cols_a = [
        ("Investor Name",       "INVESTOR_NAME",       "left",  None,   24),
        ("Entity Type",         "ENTITY_TYPE",         "left",  None,   18),
        ("Tax ID",              "TAX_ID",              "left",  None,   16),
        ("Jurisdiction",        "JURISDICTION",        "left",  None,   16),
        ("Inception Date",      "INCEPTION_DATE",      "left",  None,   14),
        ("Subscription Date",   "SUB_DATE",            "left",  None,   14),
        ("1st Contribution",    "FIRST_CONTRIB_DATE",  "left",  None,   14),
        ("Currency",            "REPT_CCY",            "left",  None,   10),
        ("Mgmt. Fee %",         "MGMT_FEE_RATE",       "right", _PCT_R, 12),
        ("Incentive Fee %",     "INC_FEE_RATE",        "right", _PCT_R, 14),
        ("Pref. Return %",      "PREF_RET",            "right", _PCT_R, 14),
        ("Hurdle Type",         "HURDLE_TYPE",         "left",  None,   14),
        ("Catch-up %",          "CATCHUP_PCT",         "right", _PCT_R, 12),
        ("Side Letter",         "SIDE_LETTER_FLG",     "left",  None,   12),
        ("FATCA",               "FATCA",               "left",  None,   14),
        ("AML/KYC",             "AML_KYC",             "left",  None,   14),
        ("Accredited",          "ACCREDITED",          "left",  None,   14),
        ("Custodian",           "CUSTODIAN",           "left",  None,   22),
        ("Special Terms",       "SPECIAL_TERMS",       "left",  None,   22),
        ("Lock-up (months)",    "LOCKUP_MO",           "right", _INT,   14),
        ("Lock-up Expired?",    "LOCKUP_EXPIRED",      "left",  None,   14),
        ("Redemp. Frequency",   "REDEMP_FREQ",         "left",  None,   16),
        ("Gate Provision",      "GATE_PROV",           "left",  None,   18),
        ("Notice (days)",       "NOTICE_DAYS",         "right", _INT,   13),
        ("DRIP Enrolled?",      "DRIP_ENROLLED",       "left",  None,   12),
        ("Side Pocket?",        "SIDE_POCKET",         "left",  None,   12),
    ]

    for ci, (label, _, align, fmt, width) in enumerate(cols_a, 1):
        _hdr(ws, r, ci, label, width=width)
    r += 1

    for i, (_, row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2
        for ci, (_, field, align, fmt, _w) in enumerate(cols_a, 1):
            try:
                v = float(_gv(row, field, 0.0)) if fmt in (_PCT_R, _INT) else _gt(row, field)
            except Exception:
                v = _gt(row, field)
            _cell(ws, r, ci, v, fmt=fmt, align=align, fill_hex=fill)
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4: PCAP
# ═══════════════════════════════════════════════════════════════════════════════

def _build_pcap(ws, pcap: pd.DataFrame) -> None:
    ws.title = "PCAP"
    r = _banner(ws, "PCAP — Full Pre-Calculated Data",
                "Complete investor PCAP as loaded from source Excel (all internal column names)", n_cols=16)

    cols = list(pcap.columns)
    for ci, col in enumerate(cols, 1):
        _hdr(ws, r, ci, col, width=16)
    r += 1

    for ri, (_, row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if ri % 2 else _GRAY2
        for ci, col in enumerate(cols, 1):
            val = _gv(row, col)
            if isinstance(val, float):
                _cell(ws, r, ci, val, fill_hex=fill)
            else:
                _cell(ws, r, ci, str(val) if val is not None else "",
                      align="left", fill_hex=fill)
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 5: Capital_Accounts
# ═══════════════════════════════════════════════════════════════════════════════

def _build_capital_accounts(ws, pcap: pd.DataFrame) -> None:
    ws.title = "Capital_Accounts"
    r = _banner(ws, "Capital Accounts — 3-Period View (CQ / YTD / ITD)", n_cols=4)

    _hdr(ws, r, 1, "Line Item",  width=38)
    _hdr(ws, r, 2, "CQ ($)",     width=20)
    _hdr(ws, r, 3, "YTD ($)",    width=20)
    _hdr(ws, r, 4, "ITD ($)",    width=20)
    r += 1

    lines = [
        ("Beginning Partner's Capital",   "BEG_CAP_CQ",  "BEG_CAP_YTD",  "BEG_CAP_ITD",  False),
        ("(+) Capital Contributions",     "CONTRIB_CQ",  "CONTRIB_YTD",  "CONTRIB_ITD",  False),
        ("(+) DRIP Reinvestment",         "DRIP_CQ",     "DRIP_YTD",     "DRIP_ITD",     False),
        ("(+) Investment Income (Loss)",  "INC_CQ",      "INC_YTD",      "INC_ITD",      False),
        ("(+) Net Unrealized Gain (Loss)","UNRLZ_CQ",    "UNRLZ_YTD",    "UNRLZ_ITD",    False),
        ("(+) Net Realized Gain (Loss)",  "RLZD_CQ",     "RLZD_YTD",     "RLZD_ITD",     False),
        ("(–) Distributions to LP",       "DIST_LP_CQ",  "DIST_LP_YTD",  "DIST_LP_ITD",  False),
        ("(–) Incentive Fees",            "INC_FEE_CQ",  "INC_FEE_YTD",  "INC_FEE_ITD",  False),
        ("Ending Partner's Capital",      "END_CAP_CQ",  "END_CAP_YTD",  "END_CAP_ITD",  True),
    ]

    for _, row in pcap.iterrows():
        _inv_banner(ws, r, _gt(row, "INVESTOR_NAME"), 4)
        r += 1
        for li, (label, cq_f, ytd_f, itd_f, bold) in enumerate(lines):
            fill = _TOTAL if bold else (_GRAY1 if li % 2 else _GRAY2)
            _cell(ws, r, 1, label,               align="left", bold=bold, fill_hex=fill)
            _cell(ws, r, 2, _gf(row, cq_f),  fmt=_USD, bold=bold, fill_hex=fill)
            _cell(ws, r, 3, _gf(row, ytd_f), fmt=_USD, bold=bold, fill_hex=fill)
            _cell(ws, r, 4, _gf(row, itd_f), fmt=_USD, bold=bold, fill_hex=fill)
            r += 1
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 6: CF_Ledger
# ═══════════════════════════════════════════════════════════════════════════════

def _build_cf_ledger(ws, cf_ledger: pd.DataFrame | None) -> None:
    ws.title = "CF_Ledger"
    r = _banner(ws, "Cashflow Ledger — Transaction-Level Data",
                "Upload a CF Ledger to populate. Required: INVESTOR_NAME, TRANSACTION_DATE, TYPE, AMOUNT", n_cols=9)

    if cf_ledger is None or cf_ledger.empty:
        _placeholder(ws, r, "No CF Ledger uploaded — provide a file with columns: "
                    "TRANSACTION_ID | INVESTOR_NAME | TRANSACTION_DATE | TYPE | AMOUNT | QUARTER | SUB_TYPE | UNITS | NOTES", 9)
        r += 1
        tmpl = ["TRANSACTION_ID", "INVESTOR_NAME", "TRANSACTION_DATE", "TYPE",
                "AMOUNT", "QUARTER", "SUB_TYPE", "UNITS", "NOTES"]
        for ci, col in enumerate(tmpl, 1):
            _hdr(ws, r, ci, col, width=18)
        return

    df = cf_ledger.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]

    for ci, col in enumerate(df.columns, 1):
        _hdr(ws, r, ci, col, width=18)
    r += 1

    for i, (_, row) in enumerate(df.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2
        for ci, col in enumerate(df.columns, 1):
            val = row[col]
            if isinstance(val, float) and pd.isna(val):
                val = ""
            _cell(ws, r, ci, val, align="left" if ci <= 4 else "right", fill_hex=fill)
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 7: CF_Aggregator
# ═══════════════════════════════════════════════════════════════════════════════

def _build_cf_aggregator(ws, pcap: pd.DataFrame, cf_ledger: pd.DataFrame | None) -> None:
    ws.title = "CF_Aggregator"
    r = _banner(ws, "CF Aggregator — Investor × Quarter Pivot",
                "Cashflow aggregation by investor and quarter", n_cols=10)

    if cf_ledger is None or cf_ledger.empty:
        _placeholder(ws, r, "Upload CF Ledger to populate this sheet", 10)
        return

    df = cf_ledger.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]

    required = {"INVESTOR_NAME", "TYPE", "AMOUNT"}
    if not required.issubset(set(df.columns)):
        _placeholder(ws, r, f"CF Ledger missing required columns: {required - set(df.columns)}", 10)
        return

    df["AMOUNT"] = pd.to_numeric(df["AMOUNT"], errors="coerce").fillna(0.0)

    if "QUARTER" not in df.columns:
        if "TRANSACTION_DATE" in df.columns:
            df["TRANSACTION_DATE"] = pd.to_datetime(df["TRANSACTION_DATE"], errors="coerce")
            df["QUARTER"] = df["TRANSACTION_DATE"].dt.to_period("Q").astype(str)
        else:
            df["QUARTER"] = "Unknown"

    records = []
    for (investor, quarter), grp in df.groupby(["INVESTOR_NAME", "QUARTER"], sort=True):
        contribs = grp.loc[grp["TYPE"] == "Contribution",  "AMOUNT"].sum()
        distribs = grp.loc[grp["TYPE"] == "Distribution",  "AMOUNT"].sum()
        drip     = grp.loc[grp["TYPE"] == "DRIP",          "AMOUNT"].sum()
        redemp   = grp.loc[grp["TYPE"] == "Redemption",    "AMOUNT"].sum()

        if "SUB_TYPE" in grp.columns:
            xfer_mask = grp["TYPE"] == "Transfer"
            xfer_in   = grp.loc[xfer_mask & (grp["SUB_TYPE"].astype(str).str.strip() == "In"),  "AMOUNT"].sum()
            xfer_out  = grp.loc[xfer_mask & (grp["SUB_TYPE"].astype(str).str.strip() == "Out"), "AMOUNT"].sum()
        else:
            xfer_in = xfer_out = 0.0

        net_cf = contribs - distribs - drip - redemp + xfer_in - xfer_out
        records.append({
            "INVESTOR_NAME": investor,
            "QUARTER":       quarter,
            "Contributions": contribs,
            "Distributions": distribs,
            "DRIP":          drip,
            "Transfers In":  xfer_in,
            "Transfers Out": xfer_out,
            "Redemptions":   redemp,
            "Net CF":        net_cf,
        })

    if not records:
        _placeholder(ws, r, "No aggregatable records in CF Ledger", 10)
        return

    hdrs   = ["Investor", "Quarter", "Contributions", "Distributions", "DRIP",
              "Transfers In", "Transfers Out", "Redemptions", "Net CF", "Running Balance"]
    widths = [26, 12, 16, 16, 14, 14, 14, 14, 16, 18]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws, r, ci, h, width=w)
    r += 1

    running: dict[str, float] = {}
    for i, rec in enumerate(records):
        inv = rec["INVESTOR_NAME"]
        running[inv] = running.get(inv, 0.0) + rec["Net CF"]
        fill = _GRAY1 if i % 2 else _GRAY2
        _cell(ws, r, 1,  rec["INVESTOR_NAME"], align="left", fill_hex=fill)
        _cell(ws, r, 2,  rec["QUARTER"],        align="left", fill_hex=fill)
        _cell(ws, r, 3,  rec["Contributions"],  fmt=_USD, fill_hex=fill)
        _cell(ws, r, 4,  rec["Distributions"],  fmt=_USD, fill_hex=fill)
        _cell(ws, r, 5,  rec["DRIP"],           fmt=_USD, fill_hex=fill)
        _cell(ws, r, 6,  rec["Transfers In"],   fmt=_USD, fill_hex=fill)
        _cell(ws, r, 7,  rec["Transfers Out"],  fmt=_USD, fill_hex=fill)
        _cell(ws, r, 8,  rec["Redemptions"],    fmt=_USD, fill_hex=fill)
        _cell(ws, r, 9,  rec["Net CF"],         fmt=_USD, fill_hex=fill)
        _cell(ws, r, 10, running[inv],          fmt=_USD, fill_hex=fill)
        r += 1

    # Totals
    _cell(ws, r, 1, "TOTAL", bold=True, align="left", fill_hex=_TOTAL)
    _cell(ws, r, 2, "",      align="left", fill_hex=_TOTAL)
    for ci, key in [(3,"Contributions"),(4,"Distributions"),(5,"DRIP"),
                    (6,"Transfers In"),(7,"Transfers Out"),(8,"Redemptions"),(9,"Net CF")]:
        total = sum(rec[key] for rec in records)
        _cell(ws, r, ci, total, fmt=_USD, bold=True, fill_hex=_TOTAL)


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 8: Distribution_Waterfall
# ═══════════════════════════════════════════════════════════════════════════════

def _build_distribution_waterfall(ws, pcap: pd.DataFrame) -> None:
    ws.title = "Distribution_Waterfall"
    r = _banner(ws, "Distribution Waterfall — ITD Per Investor",
                "Carry / waterfall mechanics: Total Return → Hurdle → Excess → GP Catch-up → LP Net", n_cols=9)

    hdrs   = ["Investor", "Contrib. ITD", "Pref. Ret %", "Hurdle Amt",
              "Total Return ITD", "Excess > Hurdle", "GP Catch-up", "LP Net WF", "WF Tier"]
    widths = [26, 18, 13, 18, 18, 18, 16, 18, 14]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws, r, ci, h, width=w)
    r += 1

    totals = {k: 0.0 for k in ["contrib","hurdle","tot_ret","excess","gp","lp"]}

    for i, (_, row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2

        contrib_itd = _gf(row, "CONTRIB_ITD")
        pref_ret    = _gf(row, "PREF_RET")
        end_cap_cq  = _gf(row, "END_CAP_CQ")
        dist_lp_itd = _gf(row, "DIST_LP_ITD")
        inc_fee_r   = _gf(row, "INC_FEE_RATE")
        catchup_pct = _gf(row, "CATCHUP_PCT", 100.0)

        # Prefer pre-computed values from PCAP; fall back to derived
        hurdle_amt  = _gf(row, "HURDLE_AMT_ITD") or (contrib_itd * pref_ret / 100.0)
        tot_ret_itd = _gf(row, "TOT_RET_ITD_DLR") or (end_cap_cq + dist_lp_itd - contrib_itd)
        excess      = _gf(row, "EXCESS_HURDLE")  if _gv(row, "EXCESS_HURDLE") is not None \
                      else max(tot_ret_itd - hurdle_amt, 0.0)
        gp_catchup  = _gf(row, "GP_CATCHUP_AMT") if _gv(row, "GP_CATCHUP_AMT") is not None \
                      else (inc_fee_r / 100.0) * (catchup_pct / 100.0) * excess
        lp_net      = _gf(row, "LP_NET_WF")      if _gv(row, "LP_NET_WF") is not None \
                      else max(excess - gp_catchup, 0.0)

        _cell(ws, r, 1, _gt(row, "INVESTOR_NAME"), align="left", fill_hex=fill)
        _cell(ws, r, 2, contrib_itd,  fmt=_USD,   fill_hex=fill)
        _cell(ws, r, 3, pref_ret,     fmt=_PCT_R, fill_hex=fill)
        _cell(ws, r, 4, hurdle_amt,   fmt=_USD,   fill_hex=fill)
        _cell(ws, r, 5, tot_ret_itd,  fmt=_USD,   fill_hex=fill)
        _cell(ws, r, 6, excess,       fmt=_USD,   fill_hex=fill)
        _cell(ws, r, 7, gp_catchup,   fmt=_USD,   fill_hex=fill)
        _cell(ws, r, 8, lp_net,       fmt=_USD,   bold=True, fill_hex=fill)
        _cell(ws, r, 9, _gt(row, "WF_TIER"), align="left", fill_hex=fill)
        r += 1

        totals["contrib"]  += contrib_itd
        totals["hurdle"]   += hurdle_amt
        totals["tot_ret"]  += tot_ret_itd
        totals["excess"]   += excess
        totals["gp"]       += gp_catchup
        totals["lp"]       += lp_net

    _cell(ws, r, 1, "TOTAL", bold=True, align="left", fill_hex=_TOTAL)
    _cell(ws, r, 2, totals["contrib"],  fmt=_USD, bold=True, fill_hex=_TOTAL)
    _cell(ws, r, 3, "",   align="left", fill_hex=_TOTAL)
    _cell(ws, r, 4, totals["hurdle"],   fmt=_USD, bold=True, fill_hex=_TOTAL)
    _cell(ws, r, 5, totals["tot_ret"],  fmt=_USD, bold=True, fill_hex=_TOTAL)
    _cell(ws, r, 6, totals["excess"],   fmt=_USD, bold=True, fill_hex=_TOTAL)
    _cell(ws, r, 7, totals["gp"],       fmt=_USD, bold=True, fill_hex=_TOTAL)
    _cell(ws, r, 8, totals["lp"],       fmt=_USD, bold=True, fill_hex=_TOTAL)


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 9: Cashflow_IRR
# ═══════════════════════════════════════════════════════════════════════════════

def _build_cashflow_irr(ws, pcap: pd.DataFrame, cf_ledger: pd.DataFrame | None) -> None:
    ws.title = "Cashflow_IRR"
    r = _banner(ws, "Cashflow IRR Analysis",
                "Gross & Net IRR cashflow vectors with XIRR formula. Values: negative=outflow, positive=inflow.", n_cols=8)

    if cf_ledger is not None and not cf_ledger.empty:
        _irr_from_ledger(ws, pcap, cf_ledger, r)
    else:
        _irr_from_pcap(ws, pcap, r)


def _irr_from_pcap(ws, pcap: pd.DataFrame, r: int) -> None:
    """2-date XIRR approximation using inception date and today."""
    today = date.today()

    # Layout:
    # Col A: Investor  B: CF_0 (inception)  C: CF_1 (terminal)  D: Date_0  E: Date_1
    # Col F: =XIRR(B:C, D:E) [Gross]   Col G: Net CF_1   Col H: =XIRR(B,G,D,E) [Net]
    # Col I: PCAP Gross IRR (check)     Col J: PCAP Net IRR (check)

    # ── Section A: Gross IRR ──────────────────────────────────────────────────
    _section_banner(ws, r, "A.  GROSS IRR — Pre-Fee Cashflows (2-date approximation)", 10)
    r += 1

    hdrs = ["Investor", "CF at Inception (−)", "CF at Terminal (+)",
            "Inception Date", "Terminal Date",
            "Gross XIRR (formula)", "PCAP Gross IRR (check)"]
    widths = [26, 22, 22, 16, 16, 22, 22]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws, r, ci, h, width=w)

    gross_data_start = r + 1

    for i, (_, row) in enumerate(pcap.iterrows()):
        r += 1
        fill = _GRAY1 if i % 2 else _GRAY2
        contrib_itd = _gf(row, "CONTRIB_ITD")
        end_cap_cq  = _gf(row, "END_CAP_CQ")
        dist_lp_itd = _gf(row, "DIST_LP_ITD")
        inc_date    = _parse_date(_gt(row, "INCEPTION_DATE")) or today
        terminal    = end_cap_cq + dist_lp_itd

        b_col = get_column_letter(2)
        c_col = get_column_letter(3)
        d_col = get_column_letter(4)
        e_col = get_column_letter(5)

        _cell(ws, r, 1, _gt(row, "INVESTOR_NAME"), align="left", fill_hex=fill)
        _cell(ws, r, 2, -contrib_itd,  fmt=_USD,    fill_hex=fill)
        _cell(ws, r, 3, terminal,      fmt=_USD,    fill_hex=fill)

        dc = ws.cell(row=r, column=4, value=inc_date)
        dc.number_format = _DATEFMT
        dc.font      = _font(size=9)
        dc.alignment = _align("left")
        dc.border    = _border()
        dc.fill      = _fill(fill)

        ec = ws.cell(row=r, column=5, value=today)
        ec.number_format = _DATEFMT
        ec.font      = _font(size=9)
        ec.alignment = _align("left")
        ec.border    = _border()
        ec.fill      = _fill(fill)

        xirr_formula = f"=XIRR({b_col}{r}:{c_col}{r},{d_col}{r}:{e_col}{r})"
        fc = ws.cell(row=r, column=6, value=xirr_formula)
        fc.number_format = "0.00%"
        fc.font      = _font(bold=True, size=9)
        fc.alignment = _align("right")
        fc.border    = _border()
        fc.fill      = _fill(_TOTAL)

        _cell(ws, r, 7, _gf(row, "GROSS_IRR") / 100.0, fmt="0.00%",
              bold=False, fill_hex=fill)

    r += 1

    # ── Section B: Net IRR ────────────────────────────────────────────────────
    r += 1
    _section_banner(ws, r, "B.  NET IRR — Post-Fee Cashflows (fees deducted from terminal value)", 10)
    r += 1

    hdrs_b = ["Investor", "CF at Inception (−)", "Net CF at Terminal (+)",
              "Inception Date", "Terminal Date",
              "Net XIRR (formula)", "PCAP Net IRR (check)"]
    for ci, (h, w) in enumerate(zip(hdrs_b, widths), 1):
        _hdr(ws, r, ci, h, width=w)

    for i, (_, row) in enumerate(pcap.iterrows()):
        r += 1
        fill = _GRAY1 if i % 2 else _GRAY2
        contrib_itd   = _gf(row, "CONTRIB_ITD")
        end_cap_cq    = _gf(row, "END_CAP_CQ")
        dist_lp_itd   = _gf(row, "DIST_LP_ITD")
        mgmt_fee_itd  = _gf(row, "MGMT_FEE_ITD_DLR")
        inc_fee_itd   = _gf(row, "INC_FEE_ITD")
        inc_date      = _parse_date(_gt(row, "INCEPTION_DATE")) or today
        net_terminal  = end_cap_cq + dist_lp_itd - mgmt_fee_itd - inc_fee_itd

        b_col = get_column_letter(2)
        c_col = get_column_letter(3)
        d_col = get_column_letter(4)
        e_col = get_column_letter(5)

        _cell(ws, r, 1, _gt(row, "INVESTOR_NAME"), align="left", fill_hex=fill)
        _cell(ws, r, 2, -contrib_itd,  fmt=_USD,   fill_hex=fill)
        _cell(ws, r, 3, net_terminal,  fmt=_USD,   fill_hex=fill)

        dc = ws.cell(row=r, column=4, value=inc_date)
        dc.number_format = _DATEFMT
        dc.font = _font(size=9); dc.alignment = _align("left")
        dc.border = _border(); dc.fill = _fill(fill)

        ec = ws.cell(row=r, column=5, value=today)
        ec.number_format = _DATEFMT
        ec.font = _font(size=9); ec.alignment = _align("left")
        ec.border = _border(); ec.fill = _fill(fill)

        xirr_formula = f"=XIRR({b_col}{r}:{c_col}{r},{d_col}{r}:{e_col}{r})"
        fc = ws.cell(row=r, column=6, value=xirr_formula)
        fc.number_format = "0.00%"
        fc.font = _font(bold=True, size=9)
        fc.alignment = _align("right")
        fc.border = _border()
        fc.fill = _fill(_TOTAL)

        _cell(ws, r, 7, _gf(row, "NET_IRR") / 100.0, fmt="0.00%", fill_hex=fill)


def _irr_from_ledger(ws, pcap: pd.DataFrame, cf_ledger: pd.DataFrame, r: int) -> None:
    """Horizontal layout: unique dates as columns, XIRR formula per investor."""
    df = cf_ledger.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]

    if "AMOUNT" not in df.columns or "INVESTOR_NAME" not in df.columns:
        _placeholder(ws, r, "CF Ledger missing INVESTOR_NAME or AMOUNT columns", 8)
        return

    df["AMOUNT"] = pd.to_numeric(df["AMOUNT"], errors="coerce").fillna(0.0)

    if "TRANSACTION_DATE" in df.columns:
        df["TRANSACTION_DATE"] = pd.to_datetime(df["TRANSACTION_DATE"], errors="coerce")
    else:
        _irr_from_pcap(ws, pcap, r)
        return

    today = date.today()

    # Sign convention: contributions = negative outflow, distributions/DRIP = positive inflow
    def _sign(txn_type: str) -> float:
        t = str(txn_type).strip()
        if t in ("Distribution", "DRIP", "Redemption"):
            return 1.0
        if t == "Contribution":
            return -1.0
        return 0.0

    # Collect all unique dates (transactions + today as terminal date)
    all_dates = sorted({d.date() for d in df["TRANSACTION_DATE"].dropna()}) + [today]

    # ── Section A: Gross IRR ──────────────────────────────────────────────────
    _section_banner(ws, r, "A.  GROSS IRR — Actual Cashflows (with terminal NAV added at report date)", len(all_dates) + 3)
    r += 1

    # Header row: dates
    _hdr(ws, r, 1, "Investor", width=26)
    date_col_start = 2
    for ci_off, dt in enumerate(all_dates):
        col_idx = date_col_start + ci_off
        c = ws.cell(row=r, column=col_idx, value=dt)
        c.number_format = _DATEFMT
        c.font      = _font(bold=True, color=_WHITE, size=9)
        c.fill      = _fill(_BLUE)
        c.alignment = _align("center")
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    irr_col = date_col_start + len(all_dates)
    _hdr(ws, r, irr_col,   "Gross XIRR",       width=18)
    _hdr(ws, r, irr_col+1, "PCAP Gross IRR",   width=18)
    hdr_row = r
    r += 1

    for i, (_, pcap_row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2
        inv_name = _gt(pcap_row, "INVESTOR_NAME")
        inv_cf = df[df["INVESTOR_NAME"] == inv_name].copy()

        cfs: dict[date, float] = {}
        for _, txn in inv_cf.iterrows():
            txn_date = txn["TRANSACTION_DATE"]
            if pd.isna(txn_date):
                continue
            d = txn_date.date()
            sign = _sign(str(txn.get("TYPE", "")))
            cfs[d] = cfs.get(d, 0.0) + sign * float(txn["AMOUNT"])

        # Add terminal NAV as final positive inflow at today
        cfs[today] = cfs.get(today, 0.0) + _gf(pcap_row, "END_CAP_CQ")

        _cell(ws, r, 1, inv_name, align="left", fill_hex=fill)
        cf_cells = []
        for ci_off, dt in enumerate(all_dates):
            col_idx = date_col_start + ci_off
            val = cfs.get(dt, 0.0)
            _cell(ws, r, col_idx, val, fmt=_USD, fill_hex=fill)
            cf_cells.append(get_column_letter(col_idx))

        # XIRR formula referencing this row's CF cells and the header row's date cells
        cf_range   = f"{cf_cells[0]}{r}:{cf_cells[-1]}{r}"
        date_range = f"{cf_cells[0]}{hdr_row}:{cf_cells[-1]}{hdr_row}"
        fc = ws.cell(row=r, column=irr_col, value=f"=XIRR({cf_range},{date_range})")
        fc.number_format = "0.00%"
        fc.font = _font(bold=True, size=9); fc.alignment = _align("right")
        fc.border = _border(); fc.fill = _fill(_TOTAL)

        _cell(ws, r, irr_col+1, _gf(pcap_row, "GROSS_IRR") / 100.0, fmt="0.00%", fill_hex=fill)
        r += 1

    # ── Section B: Net IRR ────────────────────────────────────────────────────
    r += 2
    _section_banner(ws, r, "B.  NET IRR — Cashflows Net of Fees (terminal NAV net of total fees)", len(all_dates) + 3)
    r += 1

    _hdr(ws, r, 1, "Investor", width=26)
    for ci_off, dt in enumerate(all_dates):
        col_idx = date_col_start + ci_off
        c = ws.cell(row=r, column=col_idx, value=dt)
        c.number_format = _DATEFMT
        c.font = _font(bold=True, color=_WHITE, size=9)
        c.fill = _fill(_BLUE)
        c.alignment = _align("center")
        c.border = _border()
    _hdr(ws, r, irr_col,   "Net XIRR",       width=18)
    _hdr(ws, r, irr_col+1, "PCAP Net IRR",   width=18)
    hdr_row2 = r
    r += 1

    for i, (_, pcap_row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2
        inv_name = _gt(pcap_row, "INVESTOR_NAME")
        inv_cf = df[df["INVESTOR_NAME"] == inv_name].copy()

        cfs: dict[date, float] = {}
        for _, txn in inv_cf.iterrows():
            txn_date = txn["TRANSACTION_DATE"]
            if pd.isna(txn_date):
                continue
            d = txn_date.date()
            sign = _sign(str(txn.get("TYPE", "")))
            cfs[d] = cfs.get(d, 0.0) + sign * float(txn["AMOUNT"])

        mgmt_fee = _gf(pcap_row, "MGMT_FEE_ITD_DLR")
        inc_fee  = _gf(pcap_row, "INC_FEE_ITD")
        net_nav  = _gf(pcap_row, "END_CAP_CQ") - mgmt_fee - inc_fee
        cfs[today] = cfs.get(today, 0.0) + net_nav

        _cell(ws, r, 1, inv_name, align="left", fill_hex=fill)
        cf_cells = []
        for ci_off, dt in enumerate(all_dates):
            col_idx = date_col_start + ci_off
            val = cfs.get(dt, 0.0)
            _cell(ws, r, col_idx, val, fmt=_USD, fill_hex=fill)
            cf_cells.append(get_column_letter(col_idx))

        cf_range   = f"{cf_cells[0]}{r}:{cf_cells[-1]}{r}"
        date_range = f"{cf_cells[0]}{hdr_row2}:{cf_cells[-1]}{hdr_row2}"
        fc = ws.cell(row=r, column=irr_col, value=f"=XIRR({cf_range},{date_range})")
        fc.number_format = "0.00%"
        fc.font = _font(bold=True, size=9); fc.alignment = _align("right")
        fc.border = _border(); fc.fill = _fill(_TOTAL)

        _cell(ws, r, irr_col+1, _gf(pcap_row, "NET_IRR") / 100.0, fmt="0.00%", fill_hex=fill)
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 10: Stress_IRR
# ═══════════════════════════════════════════════════════════════════════════════

def _build_stress_irr(ws, pcap: pd.DataFrame) -> None:
    ws.title = "Stress_IRR"
    r = _banner(ws, "Stress IRR — NAV Haircut Scenarios",
                "Section A: Stressed NAV | Section B: Stressed IRR | Section C: Stressed MOIC", n_cols=8)

    today = date.today()
    stress_levels = [0.0, -0.10, -0.20, -0.30, -0.15]
    stress_labels = ["Base (0%)", "−10%", "−20%", "−30%", "Custom (−15%)"]

    # ── Section A: Stressed NAV ───────────────────────────────────────────────
    _section_banner(ws, r, "A.  Stressed NAV (Ending NAV × Haircut Factor)", 7)
    r += 1

    hdrs_a   = ["Investor", "Ending NAV (Base)"] + stress_labels[1:]
    widths_a = [26, 20, 18, 18, 18, 18]
    for ci, (h, w) in enumerate(zip(hdrs_a, widths_a), 1):
        _hdr(ws, r, ci, h, width=w)
    r += 1

    nav_rows: list[list] = []
    for i, (_, row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2
        end_nav = _gf(row, "END_CAP_CQ")
        stressed = [end_nav * (1 + lvl) for lvl in stress_levels]
        nav_rows.append(stressed)
        _cell(ws, r, 1, _gt(row, "INVESTOR_NAME"), align="left", fill_hex=fill)
        for ci, val in enumerate(stressed, 2):
            _cell(ws, r, ci, val, fmt=_USD, bold=(ci == 2), fill_hex=fill)
        r += 1

    # ── Section B: Stressed IRR ───────────────────────────────────────────────
    r += 1
    _section_banner(ws, r, "B.  Stressed IRR  — (Stressed NAV + Dist ITD) / Contrib ITD)^(365/HoldDays) − 1", 7)
    r += 1

    hdrs_b = ["Investor"] + stress_labels
    for ci, (h, w) in enumerate(zip(hdrs_b, [26]+[18]*5), 1):
        _hdr(ws, r, ci, h, width=w)
    r += 1

    for i, (_, row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2
        contrib_itd = _gf(row, "CONTRIB_ITD")
        dist_lp_itd = _gf(row, "DIST_LP_ITD")
        inc_date    = _parse_date(_gt(row, "INCEPTION_DATE"))
        hold_days   = (today - inc_date).days if inc_date else 1825

        stressed_navs = nav_rows[i]
        _cell(ws, r, 1, _gt(row, "INVESTOR_NAME"), align="left", fill_hex=fill)
        for ci, (snav, lvl) in enumerate(zip(stressed_navs, stress_levels), 2):
            irr = _stress_irr(snav, dist_lp_itd, contrib_itd, hold_days)
            bold = (ci == 2)
            _cell(ws, r, ci, irr, fmt="0.00%", bold=bold, fill_hex=_TOTAL if bold else fill)
        r += 1

    # ── Section C: Stressed MOIC ──────────────────────────────────────────────
    r += 1
    _section_banner(ws, r, "C.  Stressed MOIC  — (Stressed NAV + Dist ITD) / Contrib ITD", 7)
    r += 1

    hdrs_c = ["Investor"] + stress_labels
    for ci, (h, w) in enumerate(zip(hdrs_c, [26]+[18]*5), 1):
        _hdr(ws, r, ci, h, width=w)
    r += 1

    for i, (_, row) in enumerate(pcap.iterrows()):
        fill = _GRAY1 if i % 2 else _GRAY2
        contrib_itd = _gf(row, "CONTRIB_ITD")
        dist_lp_itd = _gf(row, "DIST_LP_ITD")
        stressed_navs = nav_rows[i]
        _cell(ws, r, 1, _gt(row, "INVESTOR_NAME"), align="left", fill_hex=fill)
        for ci, snav in enumerate(stressed_navs, 2):
            moic = ((snav + dist_lp_itd) / contrib_itd) if contrib_itd > 0 else 0.0
            bold = (ci == 2)
            _cell(ws, r, ci, moic, fmt=_X, bold=bold, fill_hex=_TOTAL if bold else fill)
        r += 1

    # ── Legend ────────────────────────────────────────────────────────────────
    r += 1
    _section_banner(ws, r, "Notes & Methodology", 7)
    r += 1
    notes = [
        "Stressed IRR formula: (Stressed NAV + LP Distributions ITD) / Contrib. ITD)^(365 / Hold Days) − 1",
        "Hold Days = calendar days from Inception Date to today's date",
        "MOIC = (Stressed NAV + LP Distributions ITD) / Total Capital Contributions ITD",
        "Base (0%) = no haircut; values match PCAP Ending NAV",
        "Custom stress level set at −15%; adjust the NAV manually for other scenarios",
    ]
    for note in notes:
        ws.merge_cells(f"A{r}:H{r}")
        c = ws.cell(row=r, column=1, value=note)
        c.font = _font(italic=True, size=9)
        c.alignment = _align("left")
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Public entry point
# ═══════════════════════════════════════════════════════════════════════════════

def _build_hf_workbook_from_df(pcap_df: pd.DataFrame,
                      cf_ledger_df: pd.DataFrame | None = None) -> bytes:
    """Build the 10-sheet HF PCAP Excel workbook.

    Args:
        pcap_df:      Clean investor PCAP from hf_pcap_engine.read_hf_pcap().
        cf_ledger_df: Optional transaction-level CF ledger (INVESTOR_NAME,
                      TRANSACTION_DATE, TYPE, AMOUNT, optional: QUARTER,
                      SUB_TYPE, UNITS, NOTES).

    Returns:
        .xlsx bytes ready for st.download_button.
    """
    wb = Workbook()

    sheet_names = [
        "Period_Params", "Dashboard", "Investor_Register", "PCAP",
        "Capital_Accounts", "CF_Ledger", "CF_Aggregator",
        "Distribution_Waterfall", "Cashflow_IRR", "Stress_IRR",
    ]
    wb.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(title=name)

    _build_period_params(wb["Period_Params"],          pcap_df)
    _build_dashboard(wb["Dashboard"],                  pcap_df)
    _build_investor_register(wb["Investor_Register"],  pcap_df)
    _build_pcap(wb["PCAP"],                            pcap_df)
    _build_capital_accounts(wb["Capital_Accounts"],    pcap_df)
    _build_cf_ledger(wb["CF_Ledger"],                  cf_ledger_df)
    _build_cf_aggregator(wb["CF_Aggregator"],          pcap_df, cf_ledger_df)
    _build_distribution_waterfall(wb["Distribution_Waterfall"], pcap_df)
    _build_cashflow_irr(wb["Cashflow_IRR"],            pcap_df, cf_ledger_df)
    _build_stress_irr(wb["Stress_IRR"],                pcap_df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
# Meridian pipeline — list[dict]-based workbook builder (8 sheets)
# ══════════════════════════════════════════════════════════════════════════════

import os as _os


def _mg(inv: dict, key: str, dft: float = 0.0) -> float:
    v = inv.get(key, dft)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return dft
    try:
        return float(v)
    except (TypeError, ValueError):
        return dft


def _mgs(inv: dict, key: str, dft: str = "—") -> str:
    v = inv.get(key, dft)
    s = str(v).strip() if v is not None else ""
    return s if s and s not in ("nan", "None", "") else dft


# Meridian colour palette
_MB  = "00338D"   # KPMG Blue
_MN  = "0C233C"   # Dark Navy
_ML  = "ACEAFF"   # Light Blue
_MW  = "FFFFFF"
_MA  = "F0F4FA"   # Alt row
_MT  = "DDEEFF"   # Total row
_MS  = "E8EDF5"   # Section label
_MRD = "D73B3E"   # Red (AML flag)


def _mfill(h): return PatternFill("solid", fgColor=h)
def _mfont(bold=False, sz=10, col=_MN, italic=False):
    return Font(name="Calibri", bold=bold, size=sz, color=col, italic=italic)
def _mal(h="left", v="center", w=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=w)
def _mborder():
    s = Side(style="thin", color=_ML)
    return Border(left=s, right=s, top=s, bottom=s)


def _mhdr(ws, row: int, col: int, val, width: float | None = None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = _mfont(bold=True, sz=9, col=_MW)
    c.fill      = _mfill(_MB)
    c.alignment = _mal("center", w=True)
    c.border    = _mborder()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _mcell(ws, row: int, col: int, val=None, fmt: str | None = None,
           bold: bool = False, align: str = "right",
           fill_hex: str | None = None, red: bool = False):
    c = ws.cell(row=row, column=col, value=val)
    col_arg = _MRD if red else _MN
    c.font      = _mfont(bold=bold, sz=9, col=col_arg)
    c.alignment = _mal(align)
    c.border    = _mborder()
    if fmt:
        c.number_format = fmt
    if fill_hex:
        c.fill = _mfill(fill_hex)


# ── Sheet 1: Dashboard ────────────────────────────────────────────────────────

def _m_dashboard(ws, investors: list[dict]):
    ws.title     = "Dashboard"
    ws.tab_color = _MB

    # Title banner
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = "Meridian Opportunities Fund, L.P.  ·  Q1 2026  ·  Capital Account Dashboard"
    c.font      = _mfont(bold=True, sz=13, col=_MW)
    c.fill      = _mfill(_MB)
    c.alignment = _mal("center")
    ws.row_dimensions[1].height = 26

    # KPI row (row 3)
    ws.merge_cells("A2:N2")
    ws["A2"].value     = "Fund Key Performance Indicators"
    ws["A2"].font      = _mfont(bold=True, sz=10, col=_MN)
    ws["A2"].fill      = _mfill(_MS)
    ws["A2"].alignment = _mal("left")

    total_aum  = sum(_mg(i, "END_CAP_CQ")     for i in investors)
    total_cont = sum(_mg(i, "CONTRIB_ITD")     for i in investors)
    total_comm = sum(_mg(i, "TOTAL_COMMIT")    for i in investors)
    total_avail= sum(_mg(i, "AVAIL_COMMIT")    for i in investors)
    avg_g_irr  = sum(_mg(i, "GROSS_IRR")       for i in investors) / max(len(investors), 1)
    avg_n_irr  = sum(_mg(i, "NET_IRR")         for i in investors) / max(len(investors), 1)
    end_px     = investors[0].get("END_PX", 0) if investors else 0

    kpis = [
        ("Total AUM (CQ)",         total_aum,  _USD),
        ("Total Contributions ITD", total_cont, _USD),
        ("Total Committed",         total_comm, _USD),
        ("Available Commitment",    total_avail,_USD),
        ("# Active Investors",      len(investors), None),
        ("Ending Unit Price",       end_px,     '"$"#,##0.0000'),
        ("Fund Gross IRR",          avg_g_irr,  '0.00"%"'),
        ("Fund Net IRR",            avg_n_irr,  '0.00"%"'),
    ]
    for j, (label, val, fmt) in enumerate(kpis, 1):
        ws.cell(row=3, column=j*2-1, value=label).font = _mfont(bold=True, sz=9)
        c = ws.cell(row=3, column=j*2, value=val)
        c.font = _mfont(bold=True, sz=11, col=_MB)
        if fmt:
            c.number_format = fmt

    # Investor summary table (row 5+)
    ws.cell(row=5, column=1, value="Per-Investor Summary").font = _mfont(bold=True, sz=10)
    hdrs = ["Investor Name", "Inception Date", "Commitment", "Funded",
            "Available", "Ending Cap CQ", "Ending Cap ITD",
            "Gross IRR", "Net IRR", "% AUM", "DPI", "RVPI", "TVPI"]
    for j, h in enumerate(hdrs, 1):
        _mhdr(ws, 6, j, h, width=16 if j == 1 else 12)

    for i, inv in enumerate(investors, 7):
        aml_flag = _mgs(inv, "AML_KYC") == "In Review"
        fill_hex = _MA if i % 2 == 0 else _MW
        row_data = [
            _mgs(inv, "INVESTOR_NAME"),
            _mgs(inv, "INCEPTION_DATE"),
            _mg(inv, "TOTAL_COMMIT"),
            _mg(inv, "FUNDED_COMMIT"),
            _mg(inv, "AVAIL_COMMIT"),
            _mg(inv, "END_CAP_CQ"),
            _mg(inv, "END_CAP_ITD"),
            _mg(inv, "GROSS_IRR"),
            _mg(inv, "NET_IRR"),
            _mg(inv, "PCT_AUM"),
            _mg(inv, "DPI"),
            _mg(inv, "RVPI"),
            _mg(inv, "TVPI"),
        ]
        fmts = [None, None, _USD, _USD, _USD, _USD, _USD,
                '0.00"%"', '0.00"%"', '0.00"%"',
                '0.00"x"', '0.00"x"', '0.00"x"']
        for j, (val, fmt) in enumerate(zip(row_data, fmts), 1):
            align = "left" if j <= 2 else "right"
            _mcell(ws, i, j, val, fmt, align=align, fill_hex=fill_hex, red=aml_flag and j <= 2)

    # Totals row
    tr = 7 + len(investors)
    ws.cell(tr, 1, value="TOTAL").font = _mfont(bold=True, col=_MW)
    ws.cell(tr, 1).fill = _mfill(_MB)
    for j, col_key in enumerate([None, None, "TOTAL_COMMIT", "FUNDED_COMMIT",
                                  "AVAIL_COMMIT", "END_CAP_CQ", "END_CAP_ITD",
                                  None, None, None, None, None, None], 1):
        if col_key:
            val = sum(_mg(i, col_key) for i in investors)
            c = ws.cell(tr, j, value=val)
            c.font = _mfont(bold=True, col=_MW)
            c.fill = _mfill(_MB)
            c.number_format = _USD

    ws.freeze_panes = "A7"


# ── Sheet 2: PCAP (full data) ─────────────────────────────────────────────────

def _m_pcap(ws, investors: list[dict]):
    ws.title = "PCAP"

    if not investors:
        return
    all_keys = list(investors[0].keys())
    # Exclude non-scalar fields
    skip = {"transactions"}
    col_keys = [k for k in all_keys if k not in skip]

    for j, key in enumerate(col_keys, 1):
        _mhdr(ws, 1, j, key, width=14)

    for i, inv in enumerate(investors, 2):
        fill_hex = _MA if i % 2 == 0 else _MW
        for j, key in enumerate(col_keys, 1):
            val = inv.get(key)
            if isinstance(val, (int, float)):
                _mcell(ws, i, j, val, fill_hex=fill_hex)
            else:
                _mcell(ws, i, j, str(val) if val else "—", align="left", fill_hex=fill_hex)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(col_keys))}1"
    ws.freeze_panes = "A2"


# ── Sheet 3: Capital_Accounts ─────────────────────────────────────────────────

def _m_capital_accounts(ws, investors: list[dict]):
    ws.title = "Capital_Accounts"

    LINE_ITEMS = [
        ("Beginning Capital",             "BEG_CAP_CQ",  "BEG_CAP_YTD",  "BEG_CAP_ITD"),
        ("(+) Capital Contributions",     "CONTRIB_CQ",  "CONTRIB_YTD",  "CONTRIB_ITD"),
        ("(+) DRIP Reinvestment",         "DRIP_CQ",     "DRIP_YTD",     "DRIP_ITD"),
        ("(−) Capital Redemptions",       "REDEMP_CQ",   "REDEMP_YTD",   "REDEMP_ITD"),
        ("(+) Transfer In",               "XFER_IN_CQ",  "XFER_IN_YTD",  "XFER_IN_ITD"),
        ("(−) Transfer Out",              "XFER_OUT_CQ", "XFER_OUT_YTD", "XFER_OUT_ITD"),
        ("(+) Investment Income",         "INC_CQ",      "INC_YTD",      "INC_ITD"),
        ("(−) Fund Level Expenses",       "EXP_CQ",      "EXP_YTD",      "EXP_ITD"),
        ("(+) Net Unrealized Gain/(Loss)","UNRLZ_CQ",    "UNRLZ_YTD",    "UNRLZ_ITD"),
        ("(+) Net Realized Gain/(Loss)",  "RLZD_CQ",     "RLZD_YTD",     "RLZD_ITD"),
        ("Equity Before Distributions",   "EQ_PRED_CQ",  "EQ_PRED_YTD",  "EQ_PRED_ITD"),
        ("(−) Distributions to LP",       "DIST_LP_CQ",  "DIST_LP_YTD",  "DIST_LP_ITD"),
        ("(−) Fees to Manager",           "DIST_MGR_CQ", "DIST_MGR_YTD", "DIST_MGR_ITD"),
        ("(−) Incentive Fee",             "INC_FEE_CQ",  "INC_FEE_YTD",  "INC_FEE_ITD"),
        ("Ending Partner's Capital",      "END_CAP_CQ",  "END_CAP_YTD",  "END_CAP_ITD"),
    ]

    row = 1
    for inv in investors:
        name = _mgs(inv, "INVESTOR_NAME")
        # Investor banner
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"Investor: {name}")
        c.font = _mfont(bold=True, sz=10, col=_MW)
        c.fill = _mfill(_MN)
        ws.row_dimensions[row].height = 16
        row += 1

        # Column headers
        for j, h in enumerate(["Line Item", "CQ ($)", "YTD ($)", "ITD ($)"], 1):
            _mhdr(ws, row, j, h, width=24 if j == 1 else 14)
        row += 1

        for li, (label, cq_k, ytd_k, itd_k) in enumerate(LINE_ITEMS):
            is_total = li == len(LINE_ITEMS) - 1
            fill_hex = _MT if is_total else (_MA if li % 2 else _MW)
            _mcell(ws, row, 1, label, bold=is_total, align="left", fill_hex=fill_hex)
            for j, key in enumerate([cq_k, ytd_k, itd_k], 2):
                _mcell(ws, row, j, _mg(inv, key), fmt=_USD, bold=is_total, fill_hex=fill_hex)
            row += 1
        row += 1  # blank row between investors

    ws.freeze_panes = "A1"


# ── Sheet 4: Waterfall ────────────────────────────────────────────────────────

def _m_waterfall(ws, investors: list[dict]):
    ws.title = "Waterfall"

    hdrs = ["Investor", "Contrib ITD", "Hurdle Rate", "Hurdle Amt",
            "Mgmt Fee ITD", "Gross P&L", "Net P&L",
            "Excess Hurdle", "GP Catch-Up", "LP Pref Return",
            "LP Carry Share", "LP Net Alloc", "Ending Cap"]
    for j, h in enumerate(hdrs, 1):
        _mhdr(ws, 1, j, h, width=16 if j == 1 else 14)

    for i, inv in enumerate(investors, 2):
        aml_flag = _mgs(inv, "AML_KYC") == "In Review"
        fill_hex = _MA if i % 2 == 0 else _MW

        hurdle_r  = _mg(inv, "PREF_RET")
        gross_pnl = _mg(inv, "TOT_RET_ITD_DLR")
        mgmt_fee  = _mg(inv, "MGMT_FEE_ITD_DLR")
        net_pnl   = gross_pnl - mgmt_fee
        hurdle_a  = _mg(inv, "HURDLE_AMT_ITD")
        lp_pref   = hurdle_a
        gp_catch  = _mg(inv, "GP_CATCHUP_AMT")
        lp_carry  = _mg(inv, "EXCESS_HURDLE") * 0.80
        lp_net    = _mg(inv, "LP_NET_WF", _mg(inv, "END_CAP_ITD"))
        end_cap   = _mg(inv, "END_CAP_ITD")

        row_data = [
            _mgs(inv, "INVESTOR_NAME"),
            _mg(inv, "CONTRIB_ITD"),
            hurdle_r,
            hurdle_a,
            mgmt_fee,
            gross_pnl,
            net_pnl,
            _mg(inv, "EXCESS_HURDLE"),
            gp_catch,
            lp_pref,
            lp_carry,
            lp_net,
            end_cap,
        ]
        fmts = [None, _USD, '0.00"%"', _USD, _USD, _USD, _USD, _USD, _USD, _USD, _USD, _USD, _USD]
        for j, (val, fmt) in enumerate(zip(row_data, fmts), 1):
            align = "left" if j == 1 else "right"
            _mcell(ws, i, j, val, fmt, align=align, fill_hex=fill_hex, red=aml_flag and j == 1)

    # Totals
    tr = 2 + len(investors)
    _mcell(ws, tr, 1, "TOTAL", bold=True, align="left", fill_hex=_MB)
    ws.cell(tr, 1).font = _mfont(bold=True, col=_MW)
    for j in range(2, 14):
        col_l = get_column_letter(j)
        c = ws.cell(tr, j, value=f"=SUM({col_l}2:{col_l}{tr-1})")
        c.font         = _mfont(bold=True, col=_MW)
        c.fill         = _mfill(_MB)
        c.number_format = _USD

    ws.freeze_panes = "A2"


# ── Sheet 5: CF_Aggregator ────────────────────────────────────────────────────

def _m_cf_aggregator(ws, investors: list[dict]):
    ws.title = "CF_Aggregator"

    hdrs = ["Transaction ID", "Investor", "Date", "Quarter",
            "Type", "Sub-Type", "Amount ($)", "Units",
            "Unit Price ($)", "Status", "Notes"]
    for j, h in enumerate(hdrs, 1):
        _mhdr(ws, 1, j, h, width=16 if j in (2, 3) else 12)

    row = 2
    for inv in investors:
        for txn in inv.get("transactions", []):
            fill_hex = _MA if row % 2 == 0 else _MW
            vals = [
                txn.get("txn_id"),
                _mgs(inv, "INVESTOR_NAME"),
                txn.get("date"),
                txn.get("quarter"),
                txn.get("type"),
                txn.get("sub_type"),
                txn.get("amount"),
                txn.get("units"),
                txn.get("unit_price"),
                txn.get("status"),
                txn.get("notes"),
            ]
            fmts = [None, None, "YYYY-MM-DD", None, None, None, _USD,
                    "#,##0.0000", '"$"#,##0.0000', None, None]
            for j, (val, fmt) in enumerate(zip(vals, fmts), 1):
                align = "left" if j in (1, 2, 3, 4, 5, 6, 10, 11) else "right"
                _mcell(ws, row, j, val, fmt, align=align, fill_hex=fill_hex)
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"
    ws.freeze_panes    = "A2"


# ── Sheet 6: Cashflow_IRR ─────────────────────────────────────────────────────

def _m_cashflow_irr(ws, investors: list[dict]):
    ws.title = "Cashflow_IRR"

    all_dates: list[str] = []
    for inv in investors:
        for txn in inv.get("transactions", []):
            d = str(txn.get("date", "")) if txn.get("date") else ""
            if d and d not in all_dates:
                all_dates.append(d)
    all_dates.sort()

    if not all_dates:
        ws.cell(1, 1, "No transaction data available.").font = _mfont(italic=True)
        return

    _mhdr(ws, 1, 1, "Investor", width=22)
    for j, d in enumerate(all_dates, 2):
        _mhdr(ws, 1, j, d, width=12)
    xirr_date_col = get_column_letter(len(all_dates) + 2)
    _mhdr(ws, 1, len(all_dates) + 2, "XIRR (approx)", width=14)

    for i, inv in enumerate(investors, 2):
        fill_hex = _MA if i % 2 == 0 else _MW
        _mcell(ws, i, 1, _mgs(inv, "INVESTOR_NAME"), align="left", fill_hex=fill_hex)
        cf_by_date: dict[str, float] = {}
        for txn in inv.get("transactions", []):
            d   = str(txn.get("date", ""))
            amt = txn.get("amount")
            try:
                cf_by_date[d] = cf_by_date.get(d, 0.0) + float(amt)
            except (TypeError, ValueError):
                pass
        for j, d in enumerate(all_dates, 2):
            val = cf_by_date.get(d, 0.0)
            _mcell(ws, i, j, val if val else None, fmt=_USD, fill_hex=fill_hex)

        # XIRR formula (Excel handles this natively)
        data_start = get_column_letter(2)
        data_end   = get_column_letter(1 + len(all_dates))
        xirr_formula = (
            f"=XIRR({data_start}{i}:{data_end}{i},"
            f"{data_start}1:{data_end}1)"
        )
        c = ws.cell(i, len(all_dates) + 2, value=xirr_formula)
        c.number_format = '0.00"%"'
        c.fill = _mfill(fill_hex)
        c.font = _mfont(bold=True, sz=9)

    ws.freeze_panes = "B2"


# ── Sheet 7: Stress_IRR ───────────────────────────────────────────────────────

def _m_stress_irr(ws, investors: list[dict]):
    ws.title = "Stress_IRR"

    hdrs = ["Investor", "Base Net IRR", "NAV −10%", "NAV −20%", "NAV −30%",
            "Breakeven NAV (Pref Ret)", "Breakeven NAV (0%)"]
    for j, h in enumerate(hdrs, 1):
        _mhdr(ws, 1, j, h, width=18 if j == 1 else 16)

    for i, inv in enumerate(investors, 2):
        fill_hex  = _MA if i % 2 == 0 else _MW
        end_cap   = _mg(inv, "END_CAP_ITD")
        contrib   = _mg(inv, "CONTRIB_ITD")
        net_irr   = _mg(inv, "NET_IRR")
        pref      = _mg(inv, "PREF_RET", 8.0)

        # Approximate stressed IRR: scale terminal value by haircut, keep timing
        def _stress_irr(haircut: float) -> float | None:
            stressed_end = end_cap * (1 - haircut)
            # Approximate: IRR ≈ net_irr × (stressed_end / end_cap) if end_cap > 0
            if end_cap and contrib:
                return net_irr * (stressed_end / end_cap)
            return None

        breakeven_pref = contrib * (1 + pref / 100)
        breakeven_zero = contrib

        row_data = [
            _mgs(inv, "INVESTOR_NAME"),
            net_irr,
            _stress_irr(0.10),
            _stress_irr(0.20),
            _stress_irr(0.30),
            breakeven_pref,
            breakeven_zero,
        ]
        fmts = [None, '0.00"%"', '0.00"%"', '0.00"%"', '0.00"%"', _USD, _USD]
        for j, (val, fmt) in enumerate(zip(row_data, fmts), 1):
            align = "left" if j == 1 else "right"
            _mcell(ws, i, j, val, fmt, align=align, fill_hex=fill_hex)

    ws.freeze_panes = "A2"


# ── Sheet 8: Period_Params ────────────────────────────────────────────────────

def _m_period_params(ws, investors: list[dict]):
    ws.title = "Period_Params"

    params = [
        ("Fund Name",        "Meridian Opportunities Fund, L.P."),
        ("Report Period",    "Q1 2026"),
        ("Period Start",     "2026-01-01"),
        ("Period End",       "2026-03-31"),
        ("Ending Unit Price",investors[0].get("END_PX", 0) if investors else 0),
        ("Fund AUM",         sum(_mg(i, "END_CAP_CQ") for i in investors)),
        ("Total Committed",  sum(_mg(i, "TOTAL_COMMIT") for i in investors)),
        ("Total Investors",  len(investors)),
        ("Report Date",      date.today().isoformat()),
        ("Hurdle Rate",      "8.00% p.a."),
        ("Incentive Fee",    "20%"),
        ("Mgmt Fee",         "2.00% p.a."),
        ("Waterfall Type",   "American (deal-by-deal)"),
    ]
    _mhdr(ws, 1, 1, "Parameter", width=28)
    _mhdr(ws, 1, 2, "Value",     width=24)
    for i, (key, val) in enumerate(params, 2):
        fill_hex = _MA if i % 2 == 0 else _MW
        _mcell(ws, i, 1, key, bold=True, align="left", fill_hex=_MS)
        _mcell(ws, i, 2, val, align="left", fill_hex=fill_hex)


# ── Public entry point ────────────────────────────────────────────────────────

def _build_meridian_workbook(investors: list[dict], output_path: str) -> str:
    """
    Build the 8-sheet Meridian companion workbook.
    Saves to output_path and returns it.
    """
    wb = Workbook()

    # Create sheets in specified order
    ws_dash  = wb.active
    ws_pcap  = wb.create_sheet("PCAP")
    ws_ca    = wb.create_sheet("Capital_Accounts")
    ws_wf    = wb.create_sheet("Waterfall")
    ws_cf    = wb.create_sheet("CF_Aggregator")
    ws_irr   = wb.create_sheet("Cashflow_IRR")
    ws_str   = wb.create_sheet("Stress_IRR")
    ws_par   = wb.create_sheet("Period_Params")

    _m_dashboard(ws_dash,        investors)
    _m_pcap(ws_pcap,             investors)
    _m_capital_accounts(ws_ca,   investors)
    _m_waterfall(ws_wf,          investors)
    _m_cf_aggregator(ws_cf,      investors)
    _m_cashflow_irr(ws_irr,      investors)
    _m_stress_irr(ws_str,        investors)
    _m_period_params(ws_par,     investors)

    # Sheet tab colours
    for sheet, col in [
        (ws_pcap, _ML), (ws_ca, _ML), (ws_wf, _ML),
        (ws_cf, _ML),   (ws_irr, _ML),(ws_str, _ML),(ws_par, _ML),
    ]:
        sheet.sheet_properties.tabColor = col

    _os.makedirs(_os.path.dirname(_os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path


def build_hf_workbook(investors_or_pcap_df, output_path_or_cf_df=None,
                      output_path: str | None = None):
    """
    Dispatch wrapper — backward-compatible.

    - list[dict] (from load_pcap): saves to output_path_or_cf_df as filepath, returns str
    - pd.DataFrame (from api.py) : returns bytes (legacy path)
    """
    if isinstance(investors_or_pcap_df, list):
        return _build_meridian_workbook(investors_or_pcap_df, output_path_or_cf_df)
    return _build_hf_workbook_from_df(investors_or_pcap_df, output_path_or_cf_df)
