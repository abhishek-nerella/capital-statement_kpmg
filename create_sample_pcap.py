#!/usr/bin/env python3
"""
Creates OpenEndedFund_HedgeFund_PCAP_Q1_2026_Waterfall.xlsx
with 4 sheets: PCAP, Investor Register, CF Ledger, Waterfall.

Run once to produce the sample input file for the Meridian pipeline.
"""

from __future__ import annotations
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ───────────────────────────────────────────────────────────────────
BLUE  = "00338D"
LIGHT = "ACEAFF"
GRAY  = "F0F4FA"
WHITE = "FFFFFF"
NAVY  = "0C233C"

def _fill(c): return PatternFill("solid", fgColor=c)
def _hf(bold=False, sz=10, col=WHITE): return Font(name="Calibri", bold=bold, size=sz, color=col)
def _al(h="left", v="center", w=False): return Alignment(horizontal=h, vertical=v, wrap_text=w)


# ── PCAP column definition ─────────────────────────────────────────────────────
# (header_text, group_label, data_key)
# header_text must normalise (lowercase + collapsed whitespace) to match _ALIASES

_PCAP_COLS = [
    # Investor name
    ("Investor Name - Legal Name from Master List", "",                         "name"),
    # Unit Prices (14 cols)
    ("Beginning Partner's Capital Unit Price",       "Unit Prices",             "beg_px"),
    ("Transfer of Units Price In",                   "Unit Prices",             "xfer_in_px"),
    ("Transfer of Units Price Out",                  "Unit Prices",             "xfer_out_px"),
    ("Capital Contribution Unit Price",              "Unit Prices",             "contrib_px"),
    ("Investment and Other Income Unit Price",       "Unit Prices",             "inc_px"),
    ("Fund Level Expense Unit Price",                "Unit Prices",             "exp_px"),
    ("Net Unrealized Gain(Loss) Unit Price",         "Unit Prices",             "unrlz_px"),
    ("Net Realized Gain (Loss) Unit Price",          "Unit Prices",             "rlzd_px"),
    ("Partner's Equity Before Distributions Unit Price","Unit Prices",          "eq_pred_px"),
    ("Distr Declared to LPs Unit Price",             "Unit Prices",             "dist_lp_px"),
    ("Distr Redirected to Mgr for Fees Unit Price",  "Unit Prices",             "dist_mgr_px"),
    ("Incentive Fee Unit Price",                     "Unit Prices",             "inc_fee_px"),
    ("Reduction of Distributions for Investor Specific Taxes Unit Price",
                                                     "Unit Prices",             "tax_red_px"),
    ("Ending Partner's Capital Unit Price",          "Unit Prices",             "end_px"),
    # CQ Capital (16 cols)
    ("Beginning Partner's Capital CQ",               "CQ Capital ($)",          "beg_cap_cq"),
    ("Transfer of Units In CQ",                      "CQ Capital ($)",          "xfer_in_cq"),
    ("Transfer of Units Out CQ",                     "CQ Capital ($)",          "xfer_out_cq"),
    ("Capital Contribution CQ",                      "CQ Capital ($)",          "contrib_cq"),
    ("Contr from Div Reinvest CQ",                   "CQ Capital ($)",          "drip_cq"),
    ("Capital Redemption CQ",                        "CQ Capital ($)",          "redemp_cq"),
    ("Investment Income (Loss Before Fees) CQ",      "CQ Capital ($)",          "inc_cq"),
    ("Fund Level Expense CQ",                        "CQ Capital ($)",          "exp_cq"),
    ("Net Unrealized Gain(Loss) CQ",                 "CQ Capital ($)",          "unrlz_cq"),
    ("Net Realized Gain (Loss) CQ",                  "CQ Capital ($)",          "rlzd_cq"),
    ("Partner's Equity Before Distributions CQ",     "CQ Capital ($)",          "eq_pred_cq"),
    ("Distr Declared to LPs CQ",                     "CQ Capital ($)",          "dist_lp_cq"),
    ("Distri Redirected to Mgr for Fees CQ",         "CQ Capital ($)",          "dist_mgr_cq"),
    ("Incentive Fee CQ",                             "CQ Capital ($)",          "inc_fee_cq"),
    ("Reduction of Distributions for Investor Specific Taxes CQ",
                                                     "CQ Capital ($)",          "tax_red_cq"),
    ("Ending Partner's Capital CQ",                  "CQ Capital ($)",          "end_cap_cq"),
    # CQ Units (7 cols)
    ("Beginning Partner's Capital Units CQ",         "CQ Units",                "beg_units_cq"),
    ("Transfer Units In CQ",                         "CQ Units",                "xfer_u_in_cq"),
    ("Transfer Units Out CQ",                        "CQ Units",                "xfer_u_out_cq"),
    ("Capital Contribution Units CQ",                "CQ Units",                "contrib_u_cq"),
    ("Contr from Div Reinvest Units CQ",             "CQ Units",                "drip_u_cq"),
    ("Capital Redemption Units CQ",                  "CQ Units",                "redemp_u_cq"),
    ("Ending Partner's Capital Units CQ",            "CQ Units",                "end_units_cq"),
    # YTD Capital (16 cols)
    ("Beginning Partner's Capital YTD",              "YTD Capital ($)",         "beg_cap_ytd"),
    ("Transfer of Units In YTD",                     "YTD Capital ($)",         "xfer_in_ytd"),
    ("Transfer of Units Out YTD",                    "YTD Capital ($)",         "xfer_out_ytd"),
    ("Capital Contribution YTD",                     "YTD Capital ($)",         "contrib_ytd"),
    ("Contr from Div Reinvest YTD",                  "YTD Capital ($)",         "drip_ytd"),
    ("Capital Redemption YTD",                       "YTD Capital ($)",         "redemp_ytd"),
    ("Investment Income (Loss Before Fees) YTD",     "YTD Capital ($)",         "inc_ytd"),
    ("Fund Level Expense YTD",                       "YTD Capital ($)",         "exp_ytd"),
    ("Net Unrealized Gain(Loss) YTD",                "YTD Capital ($)",         "unrlz_ytd"),
    ("Net Realized Gain (Loss) YTD",                 "YTD Capital ($)",         "rlzd_ytd"),
    ("Partner's Capital Before Dividends YTD",       "YTD Capital ($)",         "eq_pred_ytd"),
    ("Distr Declared to LPs YTD",                    "YTD Capital ($)",         "dist_lp_ytd"),
    ("Distr Redirected to Mgr for Fees YTD",         "YTD Capital ($)",         "dist_mgr_ytd"),
    ("Incentive Fee YTD",                            "YTD Capital ($)",         "inc_fee_ytd"),
    ("Reduction of Distributions for Investor Specific Taxes YTD",
                                                     "YTD Capital ($)",         "tax_red_ytd"),
    ("Ending Partner's Capital YTD",                 "YTD Capital ($)",         "end_cap_ytd"),
    # YTD Units (7 cols)
    ("Beginning Partner's Capital Units YTD",        "YTD Units",               "beg_units_ytd"),
    ("Transfer Units In YTD",                        "YTD Units",               "xfer_u_in_ytd"),
    ("Transfer Units Out YTD",                       "YTD Units",               "xfer_u_out_ytd"),
    ("Capital Contribution Units YTD",               "YTD Units",               "contrib_u_ytd"),
    ("Contr from Div Reinvest Units YTD",            "YTD Units",               "drip_u_ytd"),
    ("Capital Redemption Units YTD",                 "YTD Units",               "redemp_u_ytd"),
    ("Ending Partner's Capital Units YTD",           "YTD Units",               "end_units_ytd"),
    # ITD Capital (16 cols)
    ("Beginning Partner's Capital ITD",              "ITD Capital ($)",         "beg_cap_itd"),
    ("Transfer of Units In ITD",                     "ITD Capital ($)",         "xfer_in_itd"),
    ("Transfer of Units Out ITD",                    "ITD Capital ($)",         "xfer_out_itd"),
    ("Capital Contribution ITD",                     "ITD Capital ($)",         "contrib_itd"),
    ("Contr from Div Reinvest ITD",                  "ITD Capital ($)",         "drip_itd"),
    ("Capital Redemption ITD",                       "ITD Capital ($)",         "redemp_itd"),
    ("Investment Income (Loss Before Fees) ITD",     "ITD Capital ($)",         "inc_itd"),
    ("Fund Level Expense ITD",                       "ITD Capital ($)",         "exp_itd"),
    ("Net Unrealized Gain(Loss) ITD",                "ITD Capital ($)",         "unrlz_itd"),
    ("Net Realized Gain (Loss) ITD",                 "ITD Capital ($)",         "rlzd_itd"),
    ("Partner's Capital Before Dividend ITD",        "ITD Capital ($)",         "eq_pred_itd"),
    ("Distr Declared to LPs ITD",                    "ITD Capital ($)",         "dist_lp_itd"),
    ("Distri Redirected to Mgr for Fee ITD",         "ITD Capital ($)",         "dist_mgr_itd"),
    ("Incentive Fees ITD",                           "ITD Capital ($)",         "inc_fee_itd"),
    ("Reduction of Distributions for Investor Specific Taxes ITD",
                                                     "ITD Capital ($)",         "tax_red_itd"),
    ("Ending Partner's Capital ITD",                 "ITD Capital ($)",         "end_cap_itd"),
    # ITD Units (7 cols)
    ("Beginning Partner's Capital Units ITD",        "ITD Units",               "beg_units_itd"),
    ("Transfer Units In ITD",                        "ITD Units",               "xfer_u_in_itd"),
    ("Transfer Units Out ITD",                       "ITD Units",               "xfer_u_out_itd"),
    ("Capital Contribution Units ITD",               "ITD Units",               "contrib_u_itd"),
    ("Contr from Div Reinvest Units ITD",            "ITD Units",               "drip_u_itd"),
    ("Capital Redemption Units ITD",                 "ITD Units",               "redemp_u_itd"),
    ("Ending Partner's Capital Units ITD",           "ITD Units",               "end_units_itd"),
    # Commitments (4 cols)
    ("Total Capital Commitment",                     "Commitments",             "total_commit"),
    ("Funded Capital Commitment",                    "Commitments",             "funded_commit"),
    ("Transfer of Commitment",                       "Commitments",             "xfer_commit"),
    ("Available Commitment",                         "Commitments",             "avail_commit"),
    # Analytics (23 cols)
    ("Gross IRR",                                    "Analytics",               "gross_irr"),
    ("Net IRR",                                      "Analytics",               "net_irr"),
    ("Investor Inception Date",                      "Analytics",               "inception_date"),
    ("% of Fund AUM",                                "Analytics",               "pct_aum"),
    ("DPI",                                          "Analytics",               "dpi"),
    ("RVPI",                                         "Analytics",               "rvpi"),
    ("TVPI",                                         "Analytics",               "tvpi"),
    ("Commitment Funded %",                          "Analytics",               "commit_funded_pct"),
    ("Unrealized Gain CQ ($)",                       "Analytics",               "unrlz_cq_dlr"),
    ("Realized Gain CQ ($)",                         "Analytics",               "rlzd_cq_dlr"),
    ("Total Return CQ ($)",                          "Analytics",               "tot_ret_cq_dlr"),
    ("Total Return CQ %",                            "Analytics",               "tot_ret_cq_pct"),
    ("Unrealized Gain ITD ($)",                      "Analytics",               "unrlz_itd_dlr"),
    ("Realized Gain ITD ($)",                        "Analytics",               "rlzd_itd_dlr"),
    ("Total Return ITD ($)",                         "Analytics",               "tot_ret_itd_dlr"),
    ("Total Return ITD %",                           "Analytics",               "tot_ret_itd_pct"),
    ("Mgmt Fee ITD ($)",                             "Analytics",               "mgmt_fee_itd_dlr"),
    ("Incentive Fee ITD ($)",                        "Analytics",               "inc_fee_itd_dlr"),
    ("Total Fees ITD ($)",                           "Analytics",               "tot_fees_itd_dlr"),
    ("Net Return ITD ($)",                           "Analytics",               "net_ret_itd_dlr"),
    ("Incentive Fee Rate",                           "Analytics",               "inc_fee_rate"),
    ("Preferred Return",                             "Analytics",               "pref_ret"),
    ("Hurdle Exceeded?",                             "Analytics",               "hurdle_exceeded"),
    # Lock-up & Redemption Terms (9 cols)
    ("Lock-up Period (Months)",                      "Lock-up & Redemption Terms", "lockup_months"),
    ("Subscription Date",                            "Lock-up & Redemption Terms", "sub_date"),
    ("First Contribution Date",                      "Lock-up & Redemption Terms", "first_contrib_date"),
    ("Redemption Eligibility Date",                  "Lock-up & Redemption Terms", "redemp_elig_date"),
    ("Lock-up Expired?",                             "Lock-up & Redemption Terms", "lockup_expired"),
    ("Redemption Frequency",                         "Lock-up & Redemption Terms", "redemp_freq"),
    ("Gate Provision",                               "Lock-up & Redemption Terms", "gate_provision"),
    ("Notice Period (Days)",                         "Lock-up & Redemption Terms", "notice_days"),
    ("Months Remaining",                             "Lock-up & Redemption Terms", "months_remaining"),
    # Side Letter Terms (11 cols)
    ("Side Pocket Eligible?",                        "Side Letter Terms",       "side_pocket"),
    ("DRIP Enrolled?",                               "Side Letter Terms",       "drip_enrolled"),
    ("Distribution Preference",                      "Side Letter Terms",       "dist_pref"),
    ("High-Water Mark Active?",                      "Side Letter Terms",       "hwm_active"),
    ("Reporting Currency",                           "Side Letter Terms",       "rept_ccy"),
    ("FATCA Status",                                 "Side Letter Terms",       "fatca"),
    ("AML/KYC Status",                               "Side Letter Terms",       "aml_kyc"),
    ("Accredited / Qualified",                       "Side Letter Terms",       "accredited"),
    ("Custodian / Prime Broker",                     "Side Letter Terms",       "custodian"),
    ("Side Letter Flag",                             "Side Letter Terms",       "side_letter"),
    ("Special Terms Notes",                          "Side Letter Terms",       "special_terms"),
    # Waterfall & Fee Parameters (11 cols)
    ("Mgmt Fee Rate",                                "Waterfall & Fee Parameters", "mgmt_fee_rate"),
    ("Incentive Fee Rate",                           "Waterfall & Fee Parameters", "inc_fee_rate2"),   # dup
    ("Preferred Return",                             "Waterfall & Fee Parameters", "pref_ret2"),       # dup
    ("Hurdle Type",                                  "Waterfall & Fee Parameters", "hurdle_type"),
    ("Catch-Up %",                                   "Waterfall & Fee Parameters", "catchup_pct"),
    ("Hurdle Amount ITD ($)",                        "Waterfall & Fee Parameters", "hurdle_amt_itd"),
    ("Total Return ITD ($)",                         "Waterfall & Fee Parameters", "tot_ret_itd2"),    # dup
    ("Excess Over Hurdle ($)",                       "Waterfall & Fee Parameters", "excess_hurdle"),
    ("GP Catch-Up Amount ($)",                       "Waterfall & Fee Parameters", "gp_catchup_amt"),
    ("LP Net Waterfall Share ($)",                   "Waterfall & Fee Parameters", "lp_net_wf"),
    ("Waterfall Tier",                               "Waterfall & Fee Parameters", "wf_tier"),
]


# ── Investor data ──────────────────────────────────────────────────────────────

def _inv(name, end_px, beg_cap, units, comm, funded,
         gross_irr, net_irr, dpi, rvpi, tvpi,
         aml, fatca, lockup, lockup_exp, months_rem,
         sub, first_c, red_elig, custodian, entity,
         pct_aum, drip_enrolled="Yes", hurdle_exceeded="Yes"):
    """Build one investor record (all numbers derived from inputs)."""
    beg_px = 1000.00
    inc_per_unit = round(end_px - beg_px + 2.50, 2)  # approx income component
    exp_per_unit = -2.50
    unrlz_per_unit = round((end_px - beg_px) * 0.65, 4)
    rlzd_per_unit  = round((end_px - beg_px) * 0.20, 4)
    inc_fee_per_unit = round(end_px - beg_px - unrlz_per_unit - rlzd_per_unit - inc_per_unit - exp_per_unit - 0.01, 4)
    eq_pred_px = round(beg_px + inc_per_unit + exp_per_unit + unrlz_per_unit + rlzd_per_unit, 4)

    # CQ $ values
    inc_cq  = round(inc_per_unit * units, 2)
    exp_cq  = round(exp_per_unit * units, 2)
    unrlz_cq = round(unrlz_per_unit * units, 2)
    rlzd_cq  = round(rlzd_per_unit  * units, 2)
    inc_fee_cq = round(inc_fee_per_unit * units, 2)
    eq_pred_cq = round(beg_cap + inc_cq + exp_cq + unrlz_cq + rlzd_cq, 2)
    end_cap_cq = round(eq_pred_cq - inc_fee_cq, 2)

    # ITD (since inception: ~4 years history)
    scale = 4.0          # 4 years of history
    contrib_itd  = funded
    drip_itd     = round(funded * 0.009, 2)
    dist_lp_itd  = round(funded * 0.12, 2)
    inc_fee_itd  = round(funded * 0.070, 2)
    inc_itd      = round(funded * scale * 0.022, 2)
    exp_itd      = round(-funded * scale * 0.0027, 2)
    unrlz_itd    = round(funded * 0.090, 2)
    rlzd_itd     = round(funded * 0.038, 2)
    eq_pred_itd  = round(contrib_itd + drip_itd + inc_itd + exp_itd + unrlz_itd + rlzd_itd, 2)
    end_cap_itd  = end_cap_cq     # ITD ending = CQ ending (current period)

    mgmt_fee_itd = round(funded * scale * 0.008, 2)
    tot_fees_itd = round(mgmt_fee_itd + inc_fee_itd, 2)
    tot_ret_itd  = round(inc_itd + exp_itd + unrlz_itd + rlzd_itd, 2)
    net_ret_itd  = round(tot_ret_itd - tot_fees_itd, 2)

    hurdle_rate = 8.00
    hurdle_amt_itd = round(funded * hurdle_rate / 100 * scale, 2)
    excess = max(0.0, round(tot_ret_itd - hurdle_amt_itd, 2))
    gp_catchup = round(excess * 0.20, 2)
    lp_net_wf  = end_cap_itd

    return {
        "name":           name,
        "beg_px":         beg_px,
        "xfer_in_px": 0, "xfer_out_px": 0,
        "contrib_px":     beg_px,
        "inc_px":         inc_per_unit,
        "exp_px":         exp_per_unit,
        "unrlz_px":       unrlz_per_unit,
        "rlzd_px":        rlzd_per_unit,
        "eq_pred_px":     eq_pred_px,
        "dist_lp_px": 0, "dist_mgr_px": 0,
        "inc_fee_px":     -abs(inc_fee_per_unit),
        "tax_red_px": 0,
        "end_px":         end_px,
        # CQ
        "beg_cap_cq": beg_cap, "xfer_in_cq": 0, "xfer_out_cq": 0,
        "contrib_cq": 0, "drip_cq": 0, "redemp_cq": 0,
        "inc_cq": inc_cq, "exp_cq": exp_cq,
        "unrlz_cq": unrlz_cq, "rlzd_cq": rlzd_cq,
        "eq_pred_cq": eq_pred_cq,
        "dist_lp_cq": 0, "dist_mgr_cq": 0,
        "inc_fee_cq": inc_fee_cq, "tax_red_cq": 0,
        "end_cap_cq": end_cap_cq,
        "beg_units_cq": units, "xfer_u_in_cq": 0, "xfer_u_out_cq": 0,
        "contrib_u_cq": 0, "drip_u_cq": 0, "redemp_u_cq": 0, "end_units_cq": units,
        # YTD = CQ (Q1 = full YTD)
        "beg_cap_ytd": beg_cap, "xfer_in_ytd": 0, "xfer_out_ytd": 0,
        "contrib_ytd": 0, "drip_ytd": 0, "redemp_ytd": 0,
        "inc_ytd": inc_cq, "exp_ytd": exp_cq,
        "unrlz_ytd": unrlz_cq, "rlzd_ytd": rlzd_cq,
        "eq_pred_ytd": eq_pred_cq,
        "dist_lp_ytd": 0, "dist_mgr_ytd": 0,
        "inc_fee_ytd": inc_fee_cq, "tax_red_ytd": 0,
        "end_cap_ytd": end_cap_cq,
        "beg_units_ytd": units, "xfer_u_in_ytd": 0, "xfer_u_out_ytd": 0,
        "contrib_u_ytd": 0, "drip_u_ytd": 0, "redemp_u_ytd": 0, "end_units_ytd": units,
        # ITD
        "beg_cap_itd": 0, "xfer_in_itd": 0, "xfer_out_itd": 0,
        "contrib_itd": contrib_itd, "drip_itd": drip_itd, "redemp_itd": 0,
        "inc_itd": inc_itd, "exp_itd": exp_itd,
        "unrlz_itd": unrlz_itd, "rlzd_itd": rlzd_itd,
        "eq_pred_itd": eq_pred_itd,
        "dist_lp_itd": dist_lp_itd, "dist_mgr_itd": 0,
        "inc_fee_itd": inc_fee_itd, "tax_red_itd": 0,
        "end_cap_itd": end_cap_itd,
        "beg_units_itd": 0, "xfer_u_in_itd": 0, "xfer_u_out_itd": 0,
        "contrib_u_itd": units, "drip_u_itd": 0, "redemp_u_itd": 0, "end_units_itd": units,
        # Commitments
        "total_commit": comm, "funded_commit": funded,
        "xfer_commit": 0, "avail_commit": round(comm - funded, 2),
        # Analytics
        "gross_irr": gross_irr, "net_irr": net_irr,
        "inception_date": first_c,
        "pct_aum": pct_aum,
        "dpi": dpi, "rvpi": rvpi, "tvpi": tvpi,
        "commit_funded_pct": round(funded / comm * 100, 2),
        "unrlz_cq_dlr": unrlz_cq, "rlzd_cq_dlr": rlzd_cq,
        "tot_ret_cq_dlr": round(inc_cq + exp_cq + unrlz_cq + rlzd_cq, 2),
        "tot_ret_cq_pct": round((end_cap_cq - beg_cap) / beg_cap * 100, 2),
        "unrlz_itd_dlr": unrlz_itd, "rlzd_itd_dlr": rlzd_itd,
        "tot_ret_itd_dlr": tot_ret_itd, "tot_ret_itd_pct": round(tvpi * 100 - 100, 2),
        "mgmt_fee_itd_dlr": mgmt_fee_itd, "inc_fee_itd_dlr": inc_fee_itd,
        "tot_fees_itd_dlr": tot_fees_itd, "net_ret_itd_dlr": net_ret_itd,
        "inc_fee_rate": 20.00, "pref_ret": hurdle_rate, "hurdle_exceeded": hurdle_exceeded,
        # Lock-up
        "lockup_months": lockup, "sub_date": sub, "first_contrib_date": first_c,
        "redemp_elig_date": red_elig, "lockup_expired": lockup_exp,
        "redemp_freq": "Quarterly", "gate_provision": "25% of NAV",
        "notice_days": 60, "months_remaining": months_rem,
        # Side letter
        "side_pocket": "No", "drip_enrolled": drip_enrolled, "dist_pref": "Standard",
        "hwm_active": "Yes", "rept_ccy": "USD", "fatca": fatca, "aml_kyc": aml,
        "accredited": "Qualified Institutional Buyer", "custodian": custodian,
        "side_letter": "No", "special_terms": "",
        # Waterfall
        "mgmt_fee_rate": 2.00, "inc_fee_rate2": 20.00, "pref_ret2": hurdle_rate,
        "hurdle_type": "Soft Hurdle", "catchup_pct": 100.00,
        "hurdle_amt_itd": hurdle_amt_itd, "tot_ret_itd2": tot_ret_itd,
        "excess_hurdle": excess, "gp_catchup_amt": gp_catchup,
        "lp_net_wf": lp_net_wf, "wf_tier": "Tier 2",
        # Register extras
        "entity_type": entity,
    }


INVESTORS = [
    _inv("Meridian Partners LP",
         end_px=1040.10, beg_cap=4_856_320.00, units=4856.32,
         comm=5_000_000, funded=4_856_320,
         gross_irr=18.52, net_irr=15.18, dpi=0.12, rvpi=1.04, tvpi=1.16,
         aml="Verified", fatca="Compliant", lockup=24, lockup_exp="Yes", months_rem=0,
         sub="01-Jan-2022", first_c="15-Jan-2022", red_elig="15-Jan-2024",
         custodian="Goldman Sachs Prime", entity="Limited Partnership", pct_aum=18.25),

    _inv("Apex Capital Management LLC",
         end_px=1042.30, beg_cap=3_121_400.00, units=3121.40,
         comm=3_500_000, funded=3_121_400,
         gross_irr=14.25, net_irr=11.80, dpi=0.10, rvpi=1.01, tvpi=1.11,
         aml="Verified", fatca="Compliant", lockup=12, lockup_exp="Yes", months_rem=0,
         sub="01-Jul-2021", first_c="01-Jul-2021", red_elig="01-Jul-2022",
         custodian="JPMorgan Prime", entity="LLC", pct_aum=11.72),

    _inv("Blue Ridge Endowment Fund",
         end_px=1038.90, beg_cap=8_432_100.00, units=8432.10,
         comm=10_000_000, funded=8_432_100,
         gross_irr=16.85, net_irr=13.92, dpi=0.14, rvpi=1.04, tvpi=1.18,
         aml="Verified", fatca="Exempt", lockup=36, lockup_exp="Yes", months_rem=0,
         sub="01-Oct-2020", first_c="01-Oct-2020", red_elig="01-Oct-2023",
         custodian="Morgan Stanley Wealth", entity="Endowment Fund", pct_aum=31.67),

    _inv("Silverstone Trust",
         end_px=1036.70, beg_cap=1_952_600.00, units=1952.60,
         comm=2_500_000, funded=1_952_600,
         gross_irr=12.48, net_irr=9.85, dpi=0.08, rvpi=1.03, tvpi=1.11,
         aml="Verified", fatca="Compliant", lockup=24, lockup_exp="No", months_rem=6,
         sub="01-Apr-2023", first_c="15-Apr-2023", red_elig="15-Apr-2025",
         custodian="UBS Prime", entity="Trust", pct_aum=7.34),

    _inv("Delta Family Office",
         end_px=1041.50, beg_cap=6_215_800.00, units=6215.80,
         comm=8_000_000, funded=6_215_800,
         gross_irr=15.92, net_irr=13.05, dpi=0.11, rvpi=1.04, tvpi=1.15,
         aml="In Review", fatca="Pending", lockup=24, lockup_exp="Yes", months_rem=0,
         sub="15-Sep-2022", first_c="01-Oct-2022", red_elig="01-Oct-2024",
         custodian="Deutsche Bank Prime", entity="Family Office", pct_aum=23.35,
         hurdle_exceeded="No"),
]


# ── Write PCAP sheet ───────────────────────────────────────────────────────────

def _write_pcap(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "PCAP"

    headers     = [c[0] for c in _PCAP_COLS]
    groups      = [c[1] for c in _PCAP_COLS]
    data_keys   = [c[2] for c in _PCAP_COLS]
    n_cols      = len(headers)

    # Row 1: group labels (merged spans)
    current_group, span_start = "", 1
    group_spans: list[tuple[int,int,str]] = []
    for j, grp in enumerate(groups, 1):
        if grp != current_group:
            if current_group and span_start < j:
                group_spans.append((span_start, j - 1, current_group))
            elif current_group:
                group_spans.append((span_start, span_start, current_group))
            current_group = grp
            span_start = j
    group_spans.append((span_start, n_cols, current_group))

    for (s, e, grp) in group_spans:
        if not grp:
            continue
        start_col = get_column_letter(s)
        end_col   = get_column_letter(e)
        if s < e:
            ws.merge_cells(f"{start_col}1:{end_col}1")
        c = ws.cell(row=1, column=s, value=grp)
        c.font      = _hf(bold=True, sz=10, col=WHITE)
        c.fill      = _fill(BLUE)
        c.alignment = _al("center")

    # Row 2: column headers
    for j, hdr in enumerate(headers, 1):
        c = ws.cell(row=2, column=j, value=hdr)
        c.font      = _hf(bold=True, sz=9, col=WHITE)
        c.fill      = _fill(NAVY)
        c.alignment = _al("center", w=True)
        ws.column_dimensions[get_column_letter(j)].width = max(12, len(hdr) * 0.85)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "B3"

    # Rows 3-7: investor data
    for i, inv in enumerate(INVESTORS, 3):
        fill_hex = GRAY if i % 2 == 0 else WHITE
        for j, key in enumerate(data_keys, 1):
            val = inv.get(key, "")
            c = ws.cell(row=i, column=j, value=val)
            c.font      = _hf(sz=9, col="000000")
            c.fill      = _fill(fill_hex)
            c.alignment = _al("right" if isinstance(val, (int, float)) else "left")


# ── Write Investor Register sheet ─────────────────────────────────────────────

def _write_register(wb: Workbook) -> None:
    ws = wb.create_sheet("Investor Register")

    hdrs = ["Investor Name (Legal)", "Entity Type", "Tax ID / EIN",
            "Jurisdiction", "Domicile", "Primary Contact", "Email", "Notes / Flags"]
    REG_MAP = {
        "Meridian Partners LP":        ("Limited Partnership",  "82-1234567", "Delaware, USA",    "United States", "John Harrison",  "jharrison@meridianpartners.com", ""),
        "Apex Capital Management LLC": ("LLC",                  "47-9876543", "New York, USA",    "United States", "Sarah Chen",     "schen@apexcap.com",              ""),
        "Blue Ridge Endowment Fund":   ("Endowment Fund",       "31-2468135", "Virginia, USA",    "United States", "Robert Miles",   "rmiles@blueridgefund.org",       ""),
        "Silverstone Trust":           ("Irrevocable Trust",    "56-1357924", "Cayman Islands",   "Cayman Islands","Emma Lockwood",  "elockwood@silverstone.ky",       ""),
        "Delta Family Office":         ("Family Office",        "63-7890123", "New York, USA",    "United States", "David Fontaine", "dfontaine@deltafo.com",          "AML/KYC under review — hold capital calls"),
    }

    # Rows 1-3: title band
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "Investor Register  ·  Meridian Opportunities Fund, L.P."
    c.font      = _hf(bold=True, sz=12, col=WHITE)
    c.fill      = _fill(BLUE)
    c.alignment = _al("center")
    ws.row_dimensions[1].height = 22

    # Row 4: headers
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font      = _hf(bold=True, sz=10, col=WHITE)
        c.fill      = _fill(NAVY)
        c.alignment = _al("center")
        ws.column_dimensions[get_column_letter(j)].width = 22

    # Rows 5-9: investor data
    for i, inv in enumerate(INVESTORS, 5):
        name = inv["name"]
        row_data = (name,) + REG_MAP.get(name, ("—","—","—","—","—","—",""))
        fill_hex = GRAY if i % 2 == 0 else WHITE
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font      = _hf(sz=9, col="000000")
            c.fill      = _fill(fill_hex)
            c.alignment = _al("left")


# ── Write CF Ledger sheet ──────────────────────────────────────────────────────

def _write_cf_ledger(wb: Workbook) -> None:
    ws = wb.create_sheet("CF Ledger")

    hdrs = ["Transaction ID", "Investor Name", "Date", "Quarter",
            "Type", "Sub-Type", "Amount", "Units", "Unit Price",
            "Running Balance", "Status", "Notes"]

    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value     = "Cash Flow Ledger  ·  Meridian Opportunities Fund, L.P."
    c.font      = _hf(bold=True, sz=12, col=WHITE)
    c.fill      = _fill(BLUE)
    c.alignment = _al("center")
    ws.row_dimensions[1].height = 22

    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font      = _hf(bold=True, sz=10, col=WHITE)
        c.fill      = _fill(NAVY)
        c.alignment = _al("center")
        ws.column_dimensions[get_column_letter(j)].width = 16

    TXNS = [
        # TX-ID, Investor, Date, Qtr, Type, SubType, Amount, Units, PX, RunBal, Status, Notes
        ("TX-001","Meridian Partners LP",        "2022-01-15","Q1 2022","Contribution","Initial", 4856320.00,4856.32,1000.00,4856320.00,"Posted","Initial capital call"),
        ("TX-002","Meridian Partners LP",        "2022-12-31","Q4 2022","Distribution","Income",  -200000.00,  -200.00,1000.00,4656320.00,"Posted","Q4 income distribution"),
        ("TX-003","Meridian Partners LP",        "2024-06-30","Q2 2024","Distribution","Return",  -380000.00,  -360.00,1055.56,5050552.30,"Posted","Capital return distribution"),
        ("TX-004","Apex Capital Management LLC", "2021-07-01","Q3 2021","Contribution","Initial", 3121400.00,3121.40,1000.00,3121400.00,"Posted","Initial capital call"),
        ("TX-005","Apex Capital Management LLC", "2023-03-31","Q1 2023","Distribution","Income",  -120000.00,  -115.00,1043.48,3001400.00,"Posted","Q1 2023 income distribution"),
        ("TX-006","Blue Ridge Endowment Fund",   "2020-10-01","Q4 2020","Contribution","Initial", 8432100.00,8432.10,1000.00,8432100.00,"Posted","Initial capital call"),
        ("TX-007","Blue Ridge Endowment Fund",   "2022-03-31","Q1 2022","Contribution","Tranche",  500000.00,  476.19,1050.00,8932100.00,"Posted","Second tranche"),
        ("TX-008","Blue Ridge Endowment Fund",   "2024-09-30","Q3 2024","Distribution","Return",  -600000.00,  -563.38,1065.38,8332100.00,"Posted","Capital return"),
        ("TX-009","Silverstone Trust",           "2023-04-15","Q2 2023","Contribution","Initial", 1952600.00,1952.60,1000.00,1952600.00,"Posted","Initial capital call"),
        ("TX-010","Delta Family Office",         "2022-10-01","Q4 2022","Contribution","Initial", 6215800.00,6215.80,1000.00,6215800.00,"Posted","Initial capital call"),
        ("TX-011","Delta Family Office",         "2023-12-31","Q4 2023","Distribution","Income",  -250000.00,  -237.53,1052.50,5965800.00,"Posted","Q4 2023 income distribution"),
        ("TX-012","Meridian Partners LP",        "2024-03-31","Q1 2024","DRIP","",                  45200.00,    43.81,1031.87,5095752.30,"Posted","DRIP reinvestment"),
        ("TX-013","Blue Ridge Endowment Fund",   "2023-06-30","Q2 2023","DRIP","",                  62400.00,    59.43,1049.93,8394500.00,"Posted","DRIP reinvestment"),
    ]

    for i, row in enumerate(TXNS, 5):
        fill_hex = GRAY if i % 2 == 0 else WHITE
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font      = _hf(sz=9, col="000000")
            c.fill      = _fill(fill_hex)
            c.alignment = _al("right" if isinstance(val, (int, float)) else "left")

    ws.auto_filter.ref = f"A4:{get_column_letter(len(hdrs))}4"
    ws.freeze_panes = "A5"


# ── Write Waterfall sheet ──────────────────────────────────────────────────────

def _write_waterfall(wb: Workbook) -> None:
    ws = wb.create_sheet("Waterfall")

    hdrs = ["Investor Name", "Capital Contrib ITD", "Hurdle Rate (%)",
            "Hurdle Amount", "Mgmt Fee (ITD)", "Gross P&L", "Net P&L",
            "Excess Over Hurdle", "GP Catch-Up", "LP Preferred Return",
            "LP Carry Share", "LP Net Allocation", "Ending Capital"]

    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value     = "Waterfall Distribution Analysis  ·  Meridian Opportunities Fund, L.P.  ·  Q1 2026"
    c.font      = _hf(bold=True, sz=12, col=WHITE)
    c.fill      = _fill(BLUE)
    c.alignment = _al("center")
    ws.row_dimensions[1].height = 22

    for j, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font      = _hf(bold=True, sz=10, col=WHITE)
        c.fill      = _fill(NAVY)
        c.alignment = _al("center", w=True)
        ws.column_dimensions[get_column_letter(j)].width = 18

    for i, inv in enumerate(INVESTORS, 5):
        ec   = inv["end_cap_itd"]
        cont = inv["contrib_itd"]
        mgmt = inv["mgmt_fee_itd_dlr"]
        tot  = inv["tot_ret_itd_dlr"]
        net  = round(tot - mgmt, 2)
        hur  = inv["hurdle_amt_itd"]
        exc  = inv["excess_hurdle"]
        gpc  = inv["gp_catchup_amt"]
        lp_p = round(hur, 2)
        lp_c = round(exc * 0.80, 2)
        lp_n = ec

        row_data = [
            inv["name"], cont, inv["pref_ret"], hur,
            mgmt, tot, net, exc, gpc, lp_p, lp_c, lp_n, ec,
        ]
        fill_hex = GRAY if i % 2 == 0 else WHITE
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font      = _hf(sz=9, col="000000")
            c.fill      = _fill(fill_hex)
            c.alignment = _al("right" if isinstance(val, (int, float)) else "left")

    # Totals row
    row = 10
    ws.cell(row=row, column=1, value="TOTAL").font = _hf(bold=True, sz=10, col=WHITE)
    ws.cell(row=row, column=1).fill = _fill(BLUE)
    for j in range(2, 14):
        ws.cell(row=row, column=j, value=f"=SUM({get_column_letter(j)}5:{get_column_letter(j)}9)")
        ws.cell(row=row, column=j).font = _hf(bold=True, sz=9, col=WHITE)
        ws.cell(row=row, column=j).fill = _fill(BLUE)

    # Step-by-step header (row 14)
    ws.cell(row=13, column=1, value="Waterfall Calculation Steps").font = _hf(bold=True, sz=10, col=BLUE)
    step_hdrs = ["Step", "Description", "Amount ($)"]
    for j, h in enumerate(step_hdrs, 1):
        c = ws.cell(row=14, column=j, value=h)
        c.font = _hf(bold=True, sz=9, col=WHITE)
        c.fill = _fill(NAVY)

    # Generic step rows (same for all investors as fund-level summary)
    steps = [
        ("1", "Gross P&L (Before Fees)",          sum(i["tot_ret_itd_dlr"]  for i in INVESTORS)),
        ("1", "Management Fee (Cost)",             -sum(i["mgmt_fee_itd_dlr"] for i in INVESTORS)),
        ("1", "Net P&L (After Mgmt Fee)",          sum(i["tot_ret_itd_dlr"] - i["mgmt_fee_itd_dlr"] for i in INVESTORS)),
        ("2", "Hurdle Amount (8% p.a.)",           sum(i["hurdle_amt_itd"]   for i in INVESTORS)),
        ("2", "LP Preferred Return",               sum(i["hurdle_amt_itd"]   for i in INVESTORS)),
        ("3", "GP Catch-Up (20%)",                 sum(i["gp_catchup_amt"]   for i in INVESTORS)),
        ("4", "LP Carry Share",                    sum(i["excess_hurdle"] * 0.80 for i in INVESTORS)),
        ("—", "LP Net Waterfall Allocation",       sum(i["lp_net_wf"]        for i in INVESTORS)),
    ]
    for i, (step, desc, amt) in enumerate(steps, 15):
        ws.cell(row=i, column=1, value=step)
        ws.cell(row=i, column=2, value=desc)
        ws.cell(row=i, column=3, value=round(amt, 2))
        fill_hex = GRAY if i % 2 == 0 else WHITE
        for j in range(1, 4):
            ws.cell(row=i, column=j).font = _hf(sz=9, col="000000")
            ws.cell(row=i, column=j).fill = _fill(fill_hex)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    out = os.path.join(os.path.dirname(__file__), "OpenEndedFund_HedgeFund_PCAP_Q1_2026_Waterfall.xlsx")
    wb  = Workbook()
    _write_pcap(wb)
    _write_register(wb)
    _write_cf_ledger(wb)
    _write_waterfall(wb)
    wb.save(out)
    print(f"✅ Created: {out}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Investors: {len(INVESTORS)}")


if __name__ == "__main__":
    main()
