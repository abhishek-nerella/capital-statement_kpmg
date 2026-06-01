"""
Hedge Fund PCAP Reader — parses the KPMG pre-calculated PCAP Excel format.

The Excel has one data row per investor. Row 1 is merged-cell group headers
("LOCK-UP & REDEMPTION TERMS", "SIDE LETTER TERMS", "WATERFALL & FEE PARAMETERS").
Row 2 is the actual column names. Data starts from row 3.

Column groups:
  Unit Prices | CQ $ | CQ Units | YTD $ | YTD Units | ITD $ | ITD Units
  | Commitments | Analytics | Lock-up & Redemption | Side Letter | Waterfall
"""

from __future__ import annotations

import re
import pandas as pd


# ── Normalisation ──────────────────────────────────────────────────────────────

def _norm(name: str) -> str:
    """Lowercase, strip, collapse all whitespace (including \\n from wrapped Excel cells)."""
    return re.sub(r"\s+", " ", str(name).replace("\n", " ")).strip().lower()


def _parse_num(val) -> float | None:
    """Parse formatted numeric strings: $1,234.56  (1,234)  1.02x  12%  -  —"""
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s in ("-", "—", "", "nan", "n/a"):
        return 0.0
    s = s.replace(",", "").replace("$", "").replace("%", "").replace("x", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


# ── Column alias table ─────────────────────────────────────────────────────────
# (internal_name, exact_normalized_column_name)
# First match wins when the same normalized name appears twice (duplicates).

_ALIASES: list[tuple[str, str]] = [
    # ── Investor identifier ────────────────────────────────────────────────────
    ("INVESTOR_NAME",       "investor name - legal name from master list"),

    # ── Unit Prices ────────────────────────────────────────────────────────────
    ("BEG_PX",              "beginning partner's capital unit price"),
    ("XFER_IN_PX",          "transfer of units price in"),
    ("XFER_OUT_PX",         "transfer of units price out"),
    ("CONTRIB_PX",          "capital contribution unit price"),
    ("INC_PX",              "investment and other income unit price"),
    ("EXP_PX",              "fund level expense unit price"),
    ("UNRLZ_PX",            "net unrealized gain(loss) unit price"),
    ("RLZD_PX",             "net realized gain (loss) unit price"),
    ("EQ_PRED_PX",          "partner's equity before distributions unit price"),
    ("DIST_LP_PX",          "distr declared to lps unit price"),
    ("DIST_MGR_PX",         "distr redirected to mgr for fees unit price"),
    ("INC_FEE_PX",          "incentive fee unit price"),
    ("TAX_RED_PX",          "reduction of distributions for investor specific taxes unit price"),
    ("END_PX",              "ending partner's capital unit price"),

    # ── CQ Capital ────────────────────────────────────────────────────────────
    ("BEG_CAP_CQ",          "beginning partner's capital cq"),
    ("XFER_IN_CQ",          "transfer of units in cq"),
    ("XFER_OUT_CQ",         "transfer of units out cq"),
    ("CONTRIB_CQ",          "capital contribution cq"),
    ("DRIP_CQ",             "contr from div reinvest cq"),
    ("REDEMP_CQ",           "capital redemption cq"),
    ("INC_CQ",              "investment income (loss before fees) cq"),
    ("EXP_CQ",              "fund level expense cq"),
    ("UNRLZ_CQ",            "net unrealized gain(loss) cq"),
    ("RLZD_CQ",             "net realized gain (loss) cq"),
    ("EQ_PRED_CQ",          "partner's equity before distributions cq"),
    ("DIST_LP_CQ",          "distr declared to lps cq"),
    ("DIST_MGR_CQ",         "distri redirected to mgr for fees cq"),
    ("INC_FEE_CQ",          "incentive fee cq"),
    ("TAX_RED_CQ",          "reduction of distributions for investor specific taxes cq"),
    ("END_CAP_CQ",          "ending partner's capital cq"),

    # ── CQ Units ──────────────────────────────────────────────────────────────
    ("BEG_UNITS_CQ",        "beginning partner's capital units cq"),
    ("XFER_U_IN_CQ",        "transfer units in cq"),
    ("XFER_U_OUT_CQ",       "transfer units out cq"),
    ("CONTRIB_U_CQ",        "capital contribution units cq"),
    ("DRIP_U_CQ",           "contr from div reinvest units cq"),
    ("REDEMP_U_CQ",         "capital redemption units cq"),
    ("END_UNITS_CQ",        "ending partner's capital units cq"),

    # ── YTD Capital ───────────────────────────────────────────────────────────
    ("BEG_CAP_YTD",         "beginning partner's capital ytd"),
    ("XFER_IN_YTD",         "transfer of units in ytd"),
    ("XFER_OUT_YTD",        "transfer of units out ytd"),
    ("CONTRIB_YTD",         "capital contribution ytd"),
    ("DRIP_YTD",            "contr from div reinvest ytd"),
    ("REDEMP_YTD",          "capital redemption ytd"),
    ("INC_YTD",             "investment income (loss before fees) ytd"),
    ("EXP_YTD",             "fund level expense ytd"),
    ("UNRLZ_YTD",           "net unrealized gain(loss) ytd"),
    ("RLZD_YTD",            "net realized gain (loss) ytd"),
    ("EQ_PRED_YTD",         "partner's capital before dividends ytd"),
    ("DIST_LP_YTD",         "distr declared to lps ytd"),
    ("DIST_MGR_YTD",        "distr redirected to mgr for fees ytd"),
    ("INC_FEE_YTD",         "incentive fee ytd"),
    ("TAX_RED_YTD",         "reduction of distributions for investor specific taxes ytd"),
    ("END_CAP_YTD",         "ending partner's capital ytd"),

    # ── YTD Units ─────────────────────────────────────────────────────────────
    ("BEG_UNITS_YTD",       "beginning partner's capital units ytd"),
    ("XFER_U_IN_YTD",       "transfer units in ytd"),
    ("XFER_U_OUT_YTD",      "transfer units out ytd"),
    ("CONTRIB_U_YTD",       "capital contribution units ytd"),
    ("DRIP_U_YTD",          "contr from div reinvest units ytd"),
    ("REDEMP_U_YTD",        "capital redemption units ytd"),
    ("END_UNITS_YTD",       "ending partner's capital units ytd"),

    # ── ITD Capital ───────────────────────────────────────────────────────────
    ("BEG_CAP_ITD",         "beginning partner's capital itd"),
    ("XFER_IN_ITD",         "transfer of units in itd"),
    ("XFER_OUT_ITD",        "transfer of units out itd"),
    ("CONTRIB_ITD",         "capital contribution itd"),
    ("DRIP_ITD",            "contr from div reinvest itd"),
    ("REDEMP_ITD",          "capital redemption itd"),
    ("INC_ITD",             "investment income (loss before fees) itd"),
    ("EXP_ITD",             "fund level expense itd"),
    ("UNRLZ_ITD",           "net unrealized gain(loss) itd"),
    ("RLZD_ITD",            "net realized gain (loss) itd"),
    ("EQ_PRED_ITD",         "partner's capital before dividend itd"),
    ("DIST_LP_ITD",         "distr declared to lps itd"),
    ("DIST_MGR_ITD",        "distri redirected to mgr for fee itd"),   # singular "fee"
    ("INC_FEE_ITD",         "incentive fees itd"),                      # plural "fees"
    ("TAX_RED_ITD",         "reduction of distributions for investor specific taxes itd"),
    ("END_CAP_ITD",         "ending partner's capital itd"),

    # ── ITD Units ─────────────────────────────────────────────────────────────
    ("BEG_UNITS_ITD",       "beginning partner's capital units itd"),
    ("XFER_U_IN_ITD",       "transfer units in itd"),
    ("XFER_U_OUT_ITD",      "transfer units out itd"),
    ("CONTRIB_U_ITD",       "capital contribution units itd"),
    ("DRIP_U_ITD",          "contr from div reinvest units itd"),
    ("REDEMP_U_ITD",        "capital redemption units itd"),
    ("END_UNITS_ITD",       "ending partner's capital units itd"),

    # ── Commitments ───────────────────────────────────────────────────────────
    ("TOTAL_COMMIT",        "total capital commitment"),
    ("FUNDED_COMMIT",       "funded capital commitment"),
    ("XFER_COMMIT",         "transfer of commitment"),
    ("AVAIL_COMMIT",        "available commitment"),

    # ── Analytics ─────────────────────────────────────────────────────────────
    ("GROSS_IRR",           "gross irr"),
    ("NET_IRR",             "net irr"),
    ("INCEPTION_DATE",      "investor inception date"),
    ("PCT_AUM",             "% of fund aum"),
    ("DPI",                 "dpi"),
    ("RVPI",                "rvpi"),
    ("TVPI",                "tvpi"),
    ("COMMIT_FUNDED_PCT",   "commitment funded %"),
    ("UNRLZ_CQ_DLR",        "unrealized gain cq ($)"),
    ("RLZD_CQ_DLR",         "realized gain cq ($)"),
    ("TOT_RET_CQ_DLR",      "total return cq ($)"),
    ("TOT_RET_CQ_PCT",      "total return cq %"),
    ("UNRLZ_ITD_DLR",       "unrealized gain itd ($)"),
    ("RLZD_ITD_DLR",        "realized gain itd ($)"),
    ("TOT_RET_ITD_DLR",     "total return itd ($)"),
    ("TOT_RET_ITD_PCT",     "total return itd %"),
    ("MGMT_FEE_ITD_DLR",    "mgmt fee itd ($)"),
    ("INC_FEE_ITD_DLR",     "incentive fee itd ($)"),
    ("TOT_FEES_ITD_DLR",    "total fees itd ($)"),
    ("NET_RET_ITD_DLR",     "net return itd ($)"),
    ("INC_FEE_RATE",        "incentive fee rate"),   # first occurrence (Analytics)
    ("PREF_RET",            "preferred return"),      # first occurrence (Analytics)
    ("HURDLE_EXCEEDED",     "hurdle exceeded?"),

    # ── Lock-up & Redemption Terms ────────────────────────────────────────────
    ("LOCKUP_MO",           "lock-up period (months)"),
    ("SUB_DATE",            "subscription date"),
    ("FIRST_CONTRIB_DATE",  "first contribution date"),
    ("REDEMP_ELIG_DATE",    "redemption eligibility date"),
    ("LOCKUP_EXPIRED",      "lock-up expired?"),
    ("REDEMP_FREQ",         "redemption frequency"),
    ("GATE_PROV",           "gate provision"),
    ("NOTICE_DAYS",         "notice period (days)"),
    ("MONTHS_REM",          "months remaining"),
    ("SIDE_POCKET",         "side pocket eligible?"),
    ("DRIP_ENROLLED",       "drip enrolled?"),
    ("DIST_PREF",           "distribution preference"),
    ("HWM_ACTIVE",          "high-water mark active?"),
    ("REPT_CCY",            "reporting currency"),
    ("FATCA",               "fatca status"),
    ("AML_KYC",             "aml/kyc status"),
    ("ACCREDITED",          "accredited / qualified"),
    ("CUSTODIAN",           "custodian / prime broker"),
    ("SIDE_LETTER_FLG",     "side letter flag"),
    ("SPECIAL_TERMS",       "special terms notes"),

    # ── Side Letter Terms ─────────────────────────────────────────────────────
    ("MGMT_FEE_RATE",       "mgmt fee rate"),
    # "incentive fee rate" and "preferred return" are duplicates of Analytics cols;
    # second occurrences will be suffixed ".1" by pandas — handled in reader.
    ("HURDLE_TYPE",         "hurdle type"),
    ("CATCHUP_PCT",         "catch-up %"),

    # ── Waterfall & Fee Parameters ────────────────────────────────────────────
    ("HURDLE_AMT_ITD",      "hurdle amount itd ($)"),
    # "total return itd ($)" is a duplicate — second occurrence skipped
    ("EXCESS_HURDLE",       "excess over hurdle ($)"),
    ("GP_CATCHUP_AMT",      "gp catch-up amount ($)"),
    ("LP_NET_WF",           "lp net waterfall share ($)"),
    ("WF_TIER",             "waterfall tier"),
]

# Primary lookup: normalized_name → internal_name (first match wins)
_NORM_TO_INTERNAL: dict[str, str] = {}
for _int, _n in _ALIASES:
    if _n not in _NORM_TO_INTERNAL:
        _NORM_TO_INTERNAL[_n] = _int

# Text-valued columns (not parsed as numeric)
_TEXT_COLS = {
    "INVESTOR_NAME", "INCEPTION_DATE", "LOCKUP_EXPIRED", "REDEMP_FREQ",
    "GATE_PROV", "SIDE_POCKET", "DRIP_ENROLLED", "DIST_PREF", "HWM_ACTIVE",
    "REPT_CCY", "FATCA", "AML_KYC", "ACCREDITED", "CUSTODIAN",
    "SIDE_LETTER_FLG", "SPECIAL_TERMS", "HURDLE_EXCEEDED", "HURDLE_TYPE",
    "WF_TIER", "SUB_DATE", "FIRST_CONTRIB_DATE", "REDEMP_ELIG_DATE",
}

# Minimum columns required for statement generation
HF_REQUIRED_COLS = {
    "ending partner's capital cq",
    "investor name - legal name from master list",
}


# ── Smart header detection ─────────────────────────────────────────────────────

def _detect_header_row(df_raw: pd.DataFrame) -> int:
    """Return the 0-based row index that contains the real column headers."""
    for i, row in df_raw.iterrows():
        for val in row:
            if isinstance(val, str) and (
                "investor name" in val.lower()
                or "legal name" in val.lower()
                or "beginning partner" in val.lower()
            ):
                return int(i)
    return 0


def read_hf_pcap_from_upload(uploaded_file) -> pd.DataFrame:
    """
    Smart reader for the KPMG HF PCAP Excel.

    Handles the two-row header pattern:
      Row 1 — merged group headers (LOCK-UP & REDEMPTION TERMS, etc.)
      Row 2 — actual column names
      Row 3+ — investor data rows
    """
    # First pass: detect which row holds the real column headers
    try:
        raw = pd.read_excel(uploaded_file, header=None, nrows=5)
        hdr_row = _detect_header_row(raw)
    except Exception:
        hdr_row = 1

    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, header=hdr_row)
    return read_hf_pcap(df)


# ── Main normalisation function ────────────────────────────────────────────────

def read_hf_pcap(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Normalise a KPMG HF PCAP DataFrame.

    1. Maps column names to short internal names.
    2. Drops blank rows and the fund totals row (no investor name).
    3. Parses formatted numeric strings → float (or keeps as str for text cols).

    Returns a clean DataFrame with one row per investor.
    """
    df = df_raw.copy()

    # Build rename map; skip columns that can't be matched
    seen_norms: set[str] = set()
    rename_map: dict[str, str] = {}

    for orig_col in df.columns:
        normed = _norm(str(orig_col))
        if normed in seen_norms:
            continue  # duplicate — keep first occurrence only
        seen_norms.add(normed)
        if normed in _NORM_TO_INTERNAL:
            rename_map[orig_col] = _NORM_TO_INTERNAL[normed]

    df.rename(columns=rename_map, inplace=True)

    # Fallback: try to locate investor name column by content if not mapped
    if "INVESTOR_NAME" not in df.columns:
        for col in df.columns:
            sample = df[col].dropna().astype(str).head(10)
            keywords = ["LP", "LLC", "Fund", "Trust", "Office", "Partners", "Capital"]
            if any(any(kw in v for kw in keywords) for v in sample):
                df.rename(columns={col: "INVESTOR_NAME"}, inplace=True)
                break

    if "INVESTOR_NAME" not in df.columns:
        raise ValueError(
            "Cannot identify investor name column. "
            "Expected 'Investor Name - Legal Name from Master List'."
        )

    # Drop rows with no investor name (blank rows, totals row)
    df["INVESTOR_NAME"] = df["INVESTOR_NAME"].astype(str).str.strip()
    df = df[
        df["INVESTOR_NAME"].notna()
        & ~df["INVESTOR_NAME"].isin(["", "nan", "None", "NaN"])
    ].copy()

    # Parse numeric columns
    for col in df.columns:
        if col in _TEXT_COLS:
            df[col] = df[col].astype(str).str.strip().replace("nan", "").replace("None", "")
        elif df[col].dtype == object:
            df[col] = df[col].apply(_parse_num)

    return df.reset_index(drop=True)


# ── Convenience accessor ───────────────────────────────────────────────────────

def _get(row: "pd.Series", field: str, default=0.0):
    """Safe field accessor — returns default if field absent or NaN."""
    val = row.get(field, default)
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    return val


def _gets(row: "pd.Series", field: str, default: str = "—") -> str:
    """Safe string accessor."""
    val = row.get(field, default)
    s = str(val).strip() if val is not None else ""
    return s if s and s not in ("nan", "None", "") else default


# ══════════════════════════════════════════════════════════════════════════════
# Meridian pipeline reader — reads the 4-sheet Meridian PCAP workbook
# ══════════════════════════════════════════════════════════════════════════════

def _openpyxl_row(ws, row_idx: int) -> list:
    """Return all cell values for a 1-indexed sheet row."""
    return [ws.cell(row=row_idx, column=j).value
            for j in range(1, ws.max_column + 1)]


def _build_col_map(headers: list) -> dict[int, str]:
    """Map 0-based column index → internal field name using _NORM_TO_INTERNAL."""
    seen: set[str] = set()
    col_map: dict[int, str] = {}
    for j, hdr in enumerate(headers):
        if hdr is None:
            continue
        normed = _norm(str(hdr))
        if normed in seen:
            continue
        seen.add(normed)
        if normed in _NORM_TO_INTERNAL:
            col_map[j] = _NORM_TO_INTERNAL[normed]
    return col_map


def _row_to_inv(headers: list, row_vals: list, col_map: dict[int, str]) -> dict:
    """Convert a raw openpyxl row into a field dict using col_map."""
    inv: dict = {}
    for col_idx, internal in col_map.items():
        if col_idx >= len(row_vals):
            continue
        val = row_vals[col_idx]
        if internal in _TEXT_COLS:
            inv[internal] = str(val).strip() if val is not None else "—"
        else:
            parsed = _parse_num(val)
            inv[internal] = parsed if parsed is not None else 0.0
    return inv


# Register field mappings: normalised header → dict key
_REG_FIELDS = {
    "investor name (legal)":   "INVESTOR_NAME",
    "entity type":             "ENTITY_TYPE",
    "tax id / ein":            "TAX_ID",
    "jurisdiction":            "JURISDICTION",
    "domicile":                "DOMICILE",
    "primary contact":         "PRIMARY_CONTACT",
    "email":                   "EMAIL",
    "notes / flags":           "NOTES_FLAGS",
}

# Waterfall sheet summary field mappings: normalised header → dict key
_WF_FIELDS = {
    "investor name":        "WF_INV_NAME",
    "capital contrib itd":  "WF_CONTRIB_ITD",
    "hurdle rate (%)":      "WF_HURDLE_RATE",
    "hurdle amount":        "WF_HURDLE_AMT",
    "mgmt fee (itd)":       "WF_MGMT_FEE",
    "gross p&l":            "WF_GROSS_PNL",
    "net p&l":              "WF_NET_PNL",
    "excess over hurdle":   "WF_EXCESS_HURDLE",
    "gp catch-up":          "WF_GP_CATCHUP",
    "lp preferred return":  "WF_LP_PREF",
    "lp carry share":       "WF_LP_CARRY",
    "lp net allocation":    "WF_LP_NET",
    "ending capital":       "WF_END_CAP",
}

# CF Ledger field mappings: normalised header → dict key
_CF_FIELDS = {
    "transaction id":   "txn_id",
    "date":             "date",
    "quarter":          "quarter",
    "type":             "type",
    "sub-type":         "sub_type",
    "amount":           "amount",
    "units":            "units",
    "unit price":       "unit_price",
    "running balance":  "running_balance",
    "status":           "status",
    "notes":            "notes",
}


def load_pcap(filepath: str) -> list[dict]:
    """
    Read OpenEndedFund_HedgeFund_PCAP_Q1_2026_Waterfall.xlsx (or any
    4-sheet PCAP workbook in the Meridian format) using openpyxl data_only=True.

    Sheets expected:
      PCAP             — row 1 = group headers, row 2 = column names, row 3+ = data
      Investor Register — row 4 = headers, row 5+ = data
      CF Ledger         — row 4 = headers, row 5+ = transactions
      Waterfall         — row 4 = headers, row 5-9 = investor summaries

    Returns a list of investor dicts.  Every PCAP field uses the internal
    short name from _ALIASES (BEG_PX, END_CAP_CQ, etc.).
    Register/Waterfall extras use the keys defined in _REG_FIELDS/_WF_FIELDS.
    CF Ledger transactions are in investor["transactions"] as list[dict].
    """
    from openpyxl import load_workbook  # import here to keep top-level import-free

    wb = load_workbook(filepath, data_only=True)

    # ── 1. PCAP sheet ──────────────────────────────────────────────────────────
    ws_pcap  = wb["PCAP"]
    pcap_hdrs = _openpyxl_row(ws_pcap, 2)   # row 2 = actual column names
    col_map   = _build_col_map(pcap_hdrs)

    investors: list[dict] = []
    for row_idx in range(3, ws_pcap.max_row + 1):
        row_vals = _openpyxl_row(ws_pcap, row_idx)
        if all(v is None for v in row_vals):
            continue
        inv = _row_to_inv(pcap_hdrs, row_vals, col_map)
        name = inv.get("INVESTOR_NAME", "")
        if not name or name in ("—", "nan", "None", ""):
            continue
        investors.append(inv)

    # ── 2. Investor Register ───────────────────────────────────────────────────
    ws_reg = wb["Investor Register"] if "Investor Register" in wb.sheetnames else None
    reg_data: dict[str, dict] = {}
    if ws_reg:
        reg_hdrs_raw = _openpyxl_row(ws_reg, 4)
        reg_col: dict[str, int] = {}  # normalised_header → 0-based col_idx
        for j, h in enumerate(reg_hdrs_raw):
            if h is not None:
                reg_col[_norm(str(h))] = j

        for row_idx in range(5, ws_reg.max_row + 1):
            row_vals = _openpyxl_row(ws_reg, row_idx)
            if all(v is None for v in row_vals):
                continue
            # Locate investor name column
            name_col = reg_col.get("investor name (legal)", reg_col.get("investor name", 0))
            raw_name = row_vals[name_col] if name_col < len(row_vals) else None
            if not raw_name:
                continue
            inv_name = str(raw_name).strip()

            entry: dict = {}
            for norm_hdr, field_key in _REG_FIELDS.items():
                col_idx = reg_col.get(norm_hdr, -1)
                if col_idx >= 0 and col_idx < len(row_vals) and row_vals[col_idx] is not None:
                    entry[field_key] = str(row_vals[col_idx]).strip()
                else:
                    # Try partial match
                    matched = next((v for k, v in {kk: row_vals[vv]
                                                    for kk, vv in reg_col.items()
                                                    if norm_hdr in kk and vv < len(row_vals)}.items()
                                    if v is not None), None)
                    entry[field_key] = str(matched).strip() if matched else "—"
            reg_data[inv_name] = entry

    for inv in investors:
        name = inv.get("INVESTOR_NAME", "")
        reg  = reg_data.get(name, {})
        for field_key in _REG_FIELDS.values():
            inv.setdefault(field_key, reg.get(field_key, "—"))

    # ── 3. CF Ledger ───────────────────────────────────────────────────────────
    ws_cf = wb["CF Ledger"] if "CF Ledger" in wb.sheetnames else None
    cf_by_investor: dict[str, list[dict]] = {}
    if ws_cf:
        cf_hdrs_raw = _openpyxl_row(ws_cf, 4)
        cf_col: dict[str, int] = {_norm(str(h)): j
                                   for j, h in enumerate(cf_hdrs_raw) if h is not None}
        inv_cf_col = cf_col.get("investor name", cf_col.get("investor", 1))

        for row_idx in range(5, ws_cf.max_row + 1):
            row_vals = _openpyxl_row(ws_cf, row_idx)
            if all(v is None for v in row_vals):
                continue
            raw = row_vals[inv_cf_col] if inv_cf_col < len(row_vals) else None
            if not raw:
                continue
            inv_name = str(raw).strip()

            txn: dict = {}
            for norm_hdr, txn_key in _CF_FIELDS.items():
                col_idx = cf_col.get(norm_hdr, -1)
                if col_idx >= 0 and col_idx < len(row_vals):
                    txn[txn_key] = row_vals[col_idx]
                else:
                    txn[txn_key] = None
            cf_by_investor.setdefault(inv_name, []).append(txn)

    for inv in investors:
        name = inv.get("INVESTOR_NAME", "")
        inv["transactions"] = cf_by_investor.get(name, [])

    # ── 4. Waterfall ───────────────────────────────────────────────────────────
    ws_wf = wb["Waterfall"] if "Waterfall" in wb.sheetnames else None
    wf_data: dict[str, dict] = {}
    if ws_wf:
        wf_hdrs_raw = _openpyxl_row(ws_wf, 4)
        wf_col: dict[str, int] = {_norm(str(h)): j
                                   for j, h in enumerate(wf_hdrs_raw) if h is not None}
        inv_wf_col = wf_col.get("investor name", 0)

        for row_idx in range(5, 10):  # rows 5-9 = 5 investors
            row_vals = _openpyxl_row(ws_wf, row_idx)
            if all(v is None for v in row_vals):
                continue
            raw = row_vals[inv_wf_col] if inv_wf_col < len(row_vals) else None
            if not raw:
                continue
            inv_name = str(raw).strip()
            entry: dict = {}
            for norm_hdr, field_key in _WF_FIELDS.items():
                if field_key == "WF_INV_NAME":
                    continue
                col_idx = wf_col.get(norm_hdr, -1)
                if col_idx >= 0 and col_idx < len(row_vals):
                    entry[field_key] = _parse_num(row_vals[col_idx]) or 0.0
                else:
                    entry[field_key] = 0.0
            wf_data[inv_name] = entry

    for inv in investors:
        name = inv.get("INVESTOR_NAME", "")
        wf = wf_data.get(name, {})
        for field_key in _WF_FIELDS.values():
            if field_key != "WF_INV_NAME":
                inv.setdefault(field_key, wf.get(field_key, 0.0))

    return investors
